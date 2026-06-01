# pyside6-uic tech_link_theme.ui -o tech_link_theme.py
# pyside6-rcc Icon.qrc -o Icon_rc.py
# pyside6-rcc icons.qrc -o icons_rc.py
# pyside6-lupdate tech_link_theme.ui -ts tech_link_theme_en.ts
# pyside6-lupdate tech_link_theme.ui -ts tech_link_theme_cn.ts

# pyinstaller --onefile --name="Packing Demo" --icon=icons\Download_Icons\robotic-arm.ico --add-binary "lib\snap7.dll;." --add-data "gifs;gifs" main.py
# pyinstaller --onefile --name="Testing App" --add-binary "lib\snap7.dll;." main_v2.py
# pyinstaller --onefile --noconsole --name="Strike Machine App" --icon=icons\strike_machine.png --add-binary "lib\snap7.dll;." --add-data "tech_link_theme_cn.qm;." main_v2.py

import sys
import os
import time 
import pandas as pd
import io
import msoffcrypto
import logging
import logging.handlers
import atexit
import sqlite3
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from PySide6.QtCore import (Qt, QTimer, QObject, QDate, QTime,
                            QSettings, QDateTime, QModelIndex,
                            QEvent, QSharedMemory, QAbstractTableModel,
                            QSystemSemaphore, QThread, QSortFilterProxyModel, 
                            Slot, QRect, QPropertyAnimation,
                            QEasingCurve, QTranslator, QLocale)
from PySide6.QtGui import (QFont, QMovie)
from PySide6.QtWidgets import (QVBoxLayout, QHeaderView, QAbstractSpinBox,
                               QLabel,QMainWindow, QApplication, QLineEdit, 
                                QFileDialog, QDialog, QTableWidget, QTableWidgetItem, QSizePolicy
)
from typing import List, Optional, Tuple, Dict, Any
from pathlib import Path
from datetime import datetime 
from openpyxl.styles import Font
from tech_link_theme import Ui_MainWindow
from Custom_Widgets import * #type: ignore
from Custom_Chart_Widgets import CustomChartWidget
# from matplot_chart_widget import CustomChartWidget
from message_box import LightThemeMessageBox as ltmessage
from password_dialog import *
from marquee_label import MarqueeLabel
from System_Data import syss
from Data_Simulator import DataSimulator
from query_plc_thread_V2 import PLCRead
from write_plc_thread_V2 import PLCWrite
from qtimer_thread import Worker

os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "0"
os.environ["QT_SCALE_FACTOR"] = "1"
os.environ["QT_FONT_DPI"] = "96"

BASE_DIR = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent

settings_path = str(BASE_DIR / "settings.ini")

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS # type: ignore
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def install_clear_on_focus(widget):
    class FocusSelectFilter(QObject):
        def eventFilter(self, obj, event):
            if event.type() == QEvent.FocusIn:
                if isinstance(widget, QAbstractSpinBox):
                    QTimer.singleShot(0, widget.lineEdit().selectAll)
                elif isinstance(widget, QLineEdit):
                    QTimer.singleShot(0, widget.selectAll)
            return False
 
    f = FocusSelectFilter(widget)
    widget.installEventFilter(f)

SIMULATE = True

class StrikeMachine(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.app_settings = QSettings(
            settings_path,
            QSettings.Format.IniFormat
        )
        # self.logger.info("INIT SETTINGS:", self.app_settings.format())
        self._find_stk_mch_folder()
        self._init_logger()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self._init_app_data()
        self._init_db_layout_and_size()
        self._init_timer()
        self._init_group_object()
        self._init_list_unit()
        self._create_charts()
        self._init_history_database()
        self._init_table_list_history()
        # self._setup_table()
        self._paint_pv_obj("#E53935")
        self._paint_sv_obj("#43A047")
        self._setup_btn_signals()
        self._setup_plc_threads(SIMULATE)
        self._translator = QTranslator()
        self.ui.home_page_btn.click()
        self.ui.clear_history_search.hide()
        self.ui.multi_search_data.hide()
        self._set_time_search_data_start_edit()
        self._set_time_search_data_end_edit()
        self.ui.stackedWidget_4.setCurrentIndex(0)

    def showEvent(self, event):# type: ignore
        super().showEvent(event)

    def _find_stk_mch_folder(self):
        # print("CURRENT SETTINGS:", self.app_settings.format())
        stk_mch_folder = Path(self.app_settings.value("stk_mch_folder", "C:\\SM_PRD", type=str))
        print("Loaded:", stk_mch_folder)

        # Nếu folder chưa tồn tại
        if not stk_mch_folder.is_dir():

            response = ltmessage.question(
                self,
                "Warning",
                "Data Folder not found! Create a new one?"
            )

            if response != ltmessage.Yes:
                return None

            parent_folder = QFileDialog.getExistingDirectory(
                self,
                "Select a Path for SM_PRD folder",
                str(stk_mch_folder.parent),
                QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks # type: ignore
            )

            if not parent_folder:
                return None

            stk_mch_folder = Path(parent_folder) / "SM_PRD"

            stk_mch_folder.mkdir(
                parents=True,
                exist_ok=True
            )

        # ALWAYS SAVE
        self.app_settings.setValue("stk_mch_folder", str(stk_mch_folder))
        self.app_settings.sync()

        self.stk_mch_folder = stk_mch_folder

    def _init_db_layout_and_size(self):
        stk_mch_file = Path(self.stk_mch_folder) / "DB_address"
        file_name = "DB Memory Address Default.xlsx"
        path = os.path.join(stk_mch_file, file_name)
        if not os.path.isfile(path):
            file_str, _ = QFileDialog.getOpenFileName(
                self,
                "Select DB Address File",
                str(stk_mch_file),
                "Excel Files (*.xlsx *.xls)"
            )
            if not file_str:  # User cancel
                return
            path = Path(file_str)

        df = self._read_protected_excel(path, password='19082002', sheet_name='Sheet1')
        if df is None:
            return

        db_layout: List[Tuple[str, str, int, Any]] = []
        max_byte = 0
        
        for i in range(5, len(df)):
            row = df.iloc[i]
            
            # Cột B: Name
            name_raw = str(row[1]).strip() if pd.notna(row[1]) else ""
            if name_raw == "" or name_raw.lower() == "nan":
                break
                
            # Chuẩn hóa tên
            name = name_raw.replace(" ", "_").replace("-", "_")
            
            # Cột C: Type
            data_type = str(row[2]).strip().upper() if pd.notna(row[2]) else ""
            
            # Cột D: Address
            try:
                addr_str = str(row[3]).strip()

                if '.' in addr_str:
                    parts = addr_str.split('.')
                    byte_addr = int(float(parts[0]))
                    bit_from_addr = int(parts[1])  # lấy bit từ cột D luôn
                else:
                    byte_addr = int(float(addr_str))
                    bit_from_addr = 0

            except:
                byte_addr = 0
                bit_from_addr = 0

            if byte_addr > max_byte:
                max_byte = byte_addr

            # Xử lý Bit cho BOOL
            bit: Any = None
            if data_type == "BOOL":
                bit = bit_from_addr  # ← lấy từ addr_str, không đọc row[4] nữa

            db_layout.append((name, data_type, byte_addr, bit))
        
        last_item = db_layout[-1] if db_layout else None
        if last_item and last_item[1] == "STRING":
            string_start = last_item[2]
            db_total_bytes = string_start + 256
        else:
            db_total_bytes = max_byte + 4
        
        self.db_dict =  {
            "ip_plc": str(df.iloc[0, 1]).strip(),
            "read_time": int(str(df.iloc[1, 1]).strip()), # iloc[collumn, row]
            "write_time": int(str(df.iloc[2, 1]).strip()), 
            "db_name": int(str(df.iloc[3, 0]).strip().replace("DB", "")),
            "DB_LAYOUT": db_layout,
            "DB_TOTAL_BYTES": db_total_bytes
        }
        # self.logger.info(f"{self.db_dict}")
        self._gui_update_connection_group(stk_mch_file)
        # print(f"Generated DB Layout: \n{self.db_dict}")

    def _read_protected_excel(self, path, password: str, sheet_name='Sheet1'):
        """Đọc file Excel có password bảo vệ."""
        try:
            with open(path, 'rb') as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                decrypted = io.BytesIO()
                office_file.decrypt(decrypted)

            decrypted.seek(0)
            return pd.read_excel(decrypted, sheet_name=sheet_name, header=None)

        except Exception as e:
            return None

    def _gui_update_connection_group(self, path_get):
        self.ui.plc_ip_address_edit.setText(self.db_dict["ip_plc"]) #type: ignore
        self.ui.db_file_path_edit.setText(str(path_get))
        self.ui.db_number_input.setValue(self.db_dict["db_name"]) #type: ignore
        self.ui.db_data_size_input.setValue(self.db_dict["DB_TOTAL_BYTES"]) #type: ignore
        self.ui.read_time_input.setValue(self.db_dict["read_time"]) #type: ignore
        self.ui.write_time_input.setValue(self.db_dict["write_time"]) #type: ignore

    def _init_logger(self):
        self.logger = logging.getLogger(__name__)
        for handler in self.logger.handlers[:]:
            self.logger.removeHandler(handler)

        # Tạo thư mục log nếu chưa có
        log_dir = self.stk_mch_folder/"Strike Machine Log"
        os.makedirs(log_dir, exist_ok=True)
        log_date = datetime.now().strftime("%d_%m_%Y")
        log_filename = os.path.join(log_dir, f'TL_SM_{log_date}.log')

        for handler in self.logger.handlers[:]:
            self.logger.removeHandler(handler)

        file_handler = logging.handlers.RotatingFileHandler( # type: ignore
            log_filename,
            maxBytes=5 * 1024 * 1024,
            backupCount=5,
            encoding='utf-8'
        )

        file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(file_formatter)
        self.logger.addHandler(file_handler)

        stream_handler = logging.StreamHandler()
        stream_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        stream_handler.setFormatter(stream_formatter)
        self.logger.addHandler(stream_handler)

        # log_handler = QPlainTextEditLogger(self)
        # gui_formatter = logging.Formatter(
        #             '%(asctime)s - %(levelname)s - %(message)s'
        #         )
        # log_handler.setFormatter(gui_formatter)
        # self.logger.addHandler(log_handler)

        self.logger.setLevel(logging.INFO)
        self.logger.propagate = False

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.MouseButtonDblClick:
            double_click_actions = {
                self.ui.plc_io_btn    : self.i_o_page_btn
            }
            if obj in double_click_actions:
                double_click_actions[obj]()
        return super().eventFilter(obj, event)

    def _show_loading_dialog(self):
        if hasattr(self, '_loading_dialog') and self._loading_dialog:
            self._loading_dialog.close()
            self._loading_dialog = None
        dialog = QDialog(self)
        dialog.setAttribute(Qt.WA_TranslucentBackground, True) # type: ignore
        dialog.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog) # type: ignore
        vbox = QVBoxLayout()
        lbl = QLabel(self)
        lbl.setStyleSheet("background: transparent;")
        gif_path = resource_path('gifs/Loading.gif')
        print(f"[Main]-[_show_loading_dialog]: Trying to load gif from: {gif_path}")
        print(f"[Main]-[_show_loading_dialog]: File exists: {os.path.exists(gif_path)}")

        self.moviee = QMovie(gif_path)
        if not self.moviee.isValid():
            print(f"[Main]-[_show_loading_dialog]: Failed to load Loading.gif from {gif_path}")
            dialog.close()
            self._loading_dialog = None
            return
        lbl.setMovie(self.moviee)
        self.moviee.start()
        vbox.addWidget(lbl)
        dialog.setLayout(vbox)
        dialog.show()
        QApplication.processEvents()
        self._loading_dialog = dialog

    def _close_loading_dialog(self):
        if hasattr(self, '_loading_dialog'):
            self._loading_dialog.close()# type: ignore
            del self._loading_dialog
        if not self._ui_shown:
            self._init_scene_before_show()
            self.center_window()
            self.show()  # Show UI chính ở đây
            # self._check_user()
            self._ui_shown = True
            QApplication.instance().installEventFilter(self) # type: ignore

    def _init_app_data(self):
        self._app = QApplication.instance()
        self._lastpos = None
        self._last_history_time = 0.0

        self._pending_rows = []
        self._all_rows_cache = []
        self._table_display = 100
        self._displayed_offset = 0
        self.history_db_path = None
        self.conn = None
        self._db_offset = 0
        self._search_keyword = ""
        self._search_offset = 0
        self._search_cache = []
        self._search_total = 0
        self._search_db_offset = 0
        self._loading_search_chunk = False
        self._search_scroll_connected = False
        self._current_unit = 0
        self._current_lang = "en"
        self._current_search_type = "name"

        self._original_data: list = []
        self._last_export_time = 0
        self.db_dict: Optional[dict] = None
        
        self.user = False
        self._dialog_open = False
        self.worker_dict = {}
        self.thread_dict = {}
        
        self._last_i_o_group_3: list = [None] * 16
        self._last_group_a:     list = [None] * 9
        self._last_group_b:     list = [None] * 9
        self._last_group_c:     list = [None] * 9
        self._last_group_a_avg: float | None = None
        self._last_group_b_avg: float | None = None
        self._last_group_c_avg: float | None = None
        self._last_t0_pv:       float | None = None
        self._last_cycle:       list = [None] * 3
        self._last_at:          float | None = None
        self._last_bt:          float | None = None
        self._last_ct:          float | None = None
        self._last_itv:         list = [None] * 3

        self.init_signal = False

        self.plc_read_worker = None
        self.plc_read_connection = False
        self.plc_read_thread = None
        self.plc_writer_worker = None
        self.plc_writer_connection = False
        self.plc_writer_thread = None

        self.default_temp_room = 25.0
        
    def _init_timer(self):
        self.all_timer = []

        self.timer_alarm = QTimer()
        self.all_timer.append(self.timer_alarm)

        self.timer_stacked_pressure_page = QTimer()
        self.all_timer.append(self.timer_stacked_pressure_page)
        
        self.chart_0_timer = QTimer()
        self.all_timer.append(self.chart_0_timer)
        self.chart_0_timer.timeout.connect(self.update_chart_temp)
        self.chart_0_timer.start(self.db_dict["read_time"])
        self.chart_1_timer = QTimer()
        self.all_timer.append(self.chart_1_timer)
        self.chart_1_timer.timeout.connect(self.update_chart_pressure_a)
        self.chart_1_timer.start(self.db_dict["read_time"])
        self.chart_2_timer = QTimer()
        self.all_timer.append(self.chart_2_timer)
        self.chart_2_timer.timeout.connect(self.update_chart_pressure_b)
        self.chart_2_timer.start(self.db_dict["read_time"])
        self.chart_3_timer = QTimer()
        self.all_timer.append(self.chart_3_timer)
        self.chart_3_timer.timeout.connect(self.update_chart_pressure_c)
        self.chart_3_timer.start(self.db_dict["read_time"])

        self._history_flush_timer = QTimer(self)
        self.all_timer.append(self._history_flush_timer)
        self._history_flush_timer.setInterval(1000)
        self._history_flush_timer.timeout.connect(self._flush_history)
        self._history_flush_timer.start()

        self.date_time_timer = QTimer(self)
        self.all_timer.append(self.date_time_timer)
        self.date_time_timer.timeout.connect(self.update_clock)
        self.date_time_timer.start(1000)
        self.update_clock()

    def _init_group_object(self):
        # self.pressure_state_obj = []
        # for i in range(10):
        #     obj = getattr(self.ui, f"pressure_sv_a_{i}")
        #     self.pressure_state_obj.append(obj)
        """
        [0] = [2]
        [1] = [3]
        [2] = [4]
        [3] = [5]
        [4] = [6]
        [5] = [7]
        [6] = [8]
        [7] = [9]
        [8] = [10]
        [9] = [11]
        [10] = [12]
        """
        self.pressure_a_pv_obj = tuple(
            getattr(self.ui, f"pressure_pv_a_{i}") 
            for i in range(2, 13)
        )
 
        self.pressure_b_pv_obj = tuple(
            getattr(self.ui, f"pressure_pv_b_{i}") 
            for i in range(2, 13)
        )

        self.pressure_c_pv_obj = tuple(
            getattr(self.ui, f"pressure_pv_c_{i}") 
            for i in range(2, 13)
        )

        skip_value = frozenset({2, 3, 4})
        """
        [0] = 1
        [1] = 5
        [2] = 6
        [3] = 7
        [4] = 8
        [5] = 9
        [6] = 10
        [7] = 11
        """
        self.pressure_a_sv_obj = tuple(
            getattr(self.ui, f"pressure_sv_a_{i}") 
            for i in range(1, 12) if i not in skip_value
        )

        self.pressure_b_sv_obj = tuple(
            getattr(self.ui, f"pressure_sv_b_{i}") 
            for i in range(1, 12) if i not in skip_value
        )

        self.pressure_c_sv_obj = tuple(
            getattr(self.ui, f"pressure_sv_c_{i}") 
            for i in range(1, 12) if i not in skip_value
        )

        self.temp_pv_obj = (
            self.ui.t0_pv,
            self.ui.at_pv,
            self.ui.bt_pv,
            self.ui.ct_pv
        )

        self.temp_sv_obj = (
            self.ui.t0_sv,
            self.ui.at_sv,
            self.ui.bt_sv,
            self.ui.ct_sv
        )

        self.temp_h_alm_obj = (
            self.ui.t0_h_alm_value,
            self.ui.at_h_alm_value,
            self.ui.bt_h_alm_value,
            self.ui.ct_h_alm_value
        )

        self.temp_l_alm_obj = (
            self.ui.t0_l_alm_value,
            self.ui.at_l_alm_value,
            self.ui.bt_l_alm_value,
            self.ui.ct_l_alm_value
        )

        self.temp_offset_obj = (
            self.ui.t0_offset_value,
            self.ui.at_t1_offset_value,
            self.ui.at_t2_offset_value,
            self.ui.at_t3_offset_value,
            self.ui.bt_t1_offset_value,
            self.ui.bt_t2_offset_value,
            self.ui.bt_t3_offset_value,
            self.ui.ct_t1_offset_value,
            self.ui.ct_t2_offset_value,
            self.ui.ct_t3_offset_value
        )

        self.list_for_import_a = (
            self.ui.pressure_sv_a_11,
            self.ui.pressure_sv_a_9,
            self.ui.pressure_sv_a_10,
            self.ui.pressure_sv_a_6,
            self.ui.pressure_sv_a_7,
            self.ui.pressure_sv_a_8,
            self.ui.pressure_sv_a_5,
            self.ui.pressure_sv_a_1,
            self.ui.at_h_alm_value,
            self.ui.at_l_alm_value,
            self.ui.at_t1_offset_value,
            self.ui.at_t2_offset_value,
            self.ui.at_t3_offset_value
        )
        self.list_for_import_b = (
            self.ui.pressure_sv_b_11,
            self.ui.pressure_sv_b_9,
            self.ui.pressure_sv_b_10,
            self.ui.pressure_sv_b_6,
            self.ui.pressure_sv_b_7,
            self.ui.pressure_sv_b_8,
            self.ui.pressure_sv_b_5,
            self.ui.pressure_sv_b_1,
            self.ui.bt_h_alm_value,
            self.ui.bt_l_alm_value,
            self.ui.bt_t1_offset_value,
            self.ui.bt_t2_offset_value,
            self.ui.bt_t3_offset_value
        )
        self.list_for_import_c = (
            self.ui.pressure_sv_c_11,
            self.ui.pressure_sv_c_9,
            self.ui.pressure_sv_c_10,
            self.ui.pressure_sv_c_6,
            self.ui.pressure_sv_c_7,
            self.ui.pressure_sv_c_8,
            self.ui.pressure_sv_c_5,
            self.ui.pressure_sv_c_1,
            self.ui.ct_h_alm_value,
            self.ui.ct_l_alm_value,
            self.ui.ct_t1_offset_value,
            self.ui.ct_t2_offset_value,
            self.ui.ct_t3_offset_value
        )
        self.list_for_import_t0 = (
            self.ui.t0_sv,
            self.ui.t0_h_alm_value,
            self.ui.t0_l_alm_value,
            self.ui.t0_offset_value
        )

        self.io_group_1_switch_obj = tuple(
            getattr(self.ui, f"i_o_group_1_switch_{i}") for i in range(2, 14)
        )
        self.i_o_group_3_obj = (
            self.ui.t0_value,
            self.ui.t1_1_value,
            self.ui.t1_2_value,
            self.ui.t1_3_value,
            self.ui.t2_1_value,
            self.ui.t2_2_value,
            self.ui.t2_3_value,
            self.ui.t3_1_value,
            self.ui.t3_2_value,
            self.ui.t3_3_value,
            self.ui.p1_value,
            self.ui.p2_value,
            self.ui.p3_value,
            self.ui.fp1_value,
            self.ui.fp2_value,
            self.ui.fp3_value
        )
        # self.all_widgets = (
        #     self.pressure_state_obj +
        #     self.inlet_temp_obj     +
        #     self.front_temp_obj     +
        #     self.middle_temp_obj    +
        #     self.end_temp_obj
        # )
        # assert len(self.all_widgets) == 22, f"Widget count mismatch: {len(self.all_widgets)}"

    def _paint_pv_obj(self, color):
        for obj in self.pressure_a_pv_obj + self.pressure_b_pv_obj + self.pressure_c_pv_obj + self.temp_pv_obj:
            obj.setStyleSheet(f"""
                QDoubleSpinBox {{
                    color: {color};}}
                QSpinBox {{
                    color: {color};}}""")
            
    def _paint_sv_obj(self, color):
        for obj in self.pressure_a_sv_obj + self.pressure_b_sv_obj + self.pressure_c_sv_obj + self.temp_offset_obj + self.temp_h_alm_obj + self.temp_l_alm_obj + self.temp_sv_obj:
            obj.setStyleSheet(f"""
                QDoubleSpinBox {{
                    color: {color};}}
                QDoubleSpinBox:hover {{
                    border: 2px solid {color};
                }}
                QSpinBox {{
                    color: {color};}}
                QSpinBox:hover {{
                    border: 2px solid {color};
                }}
            """)

    def _init_list_unit(self):
        self.cel_fah_change = [
            self.ui.stacked_cel_fah_press_a_1,
            self.ui.stacked_cel_fah_press_a_2,
            self.ui.stacked_cel_fah_press_a_3,
            self.ui.stacked_cel_fah_press_a_4,

            self.ui.stacked_cel_fah_press_b_1,
            self.ui.stacked_cel_fah_press_b_2,
            self.ui.stacked_cel_fah_press_b_3,
            self.ui.stacked_cel_fah_press_b_4,

            self.ui.stacked_cel_fah_press_c_1,
            self.ui.stacked_cel_fah_press_c_2,
            self.ui.stacked_cel_fah_press_c_3,
            self.ui.stacked_cel_fah_press_c_4,

            self.ui.stacked_cel_fah_temp_t0_1,
            self.ui.stacked_cel_fah_temp_t0_2,
            self.ui.stacked_cel_fah_temp_t0_3,
            self.ui.stacked_cel_fah_temp_t0_4,
            self.ui.stacked_cel_fah_temp_t0_5,

            self.ui.stacked_cel_fah_temp_a_1,
            self.ui.stacked_cel_fah_temp_a_2,
            self.ui.stacked_cel_fah_temp_a_3,
            self.ui.stacked_cel_fah_temp_a_4,
            self.ui.stacked_cel_fah_temp_a_5,
            self.ui.stacked_cel_fah_temp_a_6,
            self.ui.stacked_cel_fah_temp_a_7,

            self.ui.stacked_cel_fah_temp_b_1,
            self.ui.stacked_cel_fah_temp_b_2,
            self.ui.stacked_cel_fah_temp_b_3,
            self.ui.stacked_cel_fah_temp_b_4,
            self.ui.stacked_cel_fah_temp_b_5,
            self.ui.stacked_cel_fah_temp_b_6,
            self.ui.stacked_cel_fah_temp_b_7,

            self.ui.stacked_cel_fah_temp_c_1,
            self.ui.stacked_cel_fah_temp_c_2,
            self.ui.stacked_cel_fah_temp_c_3,
            self.ui.stacked_cel_fah_temp_c_4,
            self.ui.stacked_cel_fah_temp_c_5,
            self.ui.stacked_cel_fah_temp_c_6,
            self.ui.stacked_cel_fah_temp_c_7
        ]

    def _create_charts(self):
        font = QFont("Segoe UI", 17)
        font.setWeight(QFont.Weight.Bold)

        # Chart Nhiệt độ (Oven)
        self.chart_temp = CustomChartWidget(
            title="Oven",
            num_temp=2,
            num_pressure=0,
            temp_label="Temperature (°C)",
            pressure_label="",
            temp_range=(0, 50),
            pressure_range=(0, 0),
            max_seconds=60,
            chart_font=font
        )

        # Chart Group A
        self.chart_pressure_a = CustomChartWidget(
            title="Group A",
            num_temp=4,
            num_pressure=2,
            temp_label="Temperature (°C)",
            pressure_label="Pressure (bar)",
            temp_range=(0, 50),
            pressure_range=(0, 15),
            max_seconds=60,
            chart_font=font
        )

        # Chart Group B
        self.chart_pressure_b = CustomChartWidget(
            title="Group B",
            num_temp=4,
            num_pressure=2,
            temp_label="Temperature (°C)",
            pressure_label="Pressure (bar)",
            temp_range=(0, 50),
            pressure_range=(0, 15),
            max_seconds=60,
            chart_font=font
        )

        # Chart Group C
        self.chart_pressure_c = CustomChartWidget(
            title="Group C",
            num_temp=4,
            num_pressure=2,
            temp_label="Temperature (°C)",
            pressure_label="Pressure (bar)",
            temp_range=(0, 50),
            pressure_range=(0, 15),
            max_seconds=60,
            chart_font=font
        )

        # Kết nối nút tiêu đề
        self.chart_temp.btn_setting.clicked.connect(self.temperature_page_btn)
        self.chart_pressure_a.btn_setting.clicked.connect(self.ui.home_page_btn.click)
        self.chart_pressure_b.btn_setting.clicked.connect(self.ui.home_page_btn.click)
        self.chart_pressure_c.btn_setting.clicked.connect(self.ui.home_page_btn.click)

        # Thêm vào layout
        self.ui.card_temperature.addWidget(self.chart_temp)
        self.ui.card_pressure_1.addWidget(self.chart_pressure_a)
        self.ui.card_pressure_2.addWidget(self.chart_pressure_b)
        self.ui.card_pressure_3.addWidget(self.chart_pressure_c)
        self._chart_frames = [
                self.ui.card_temperature.parentWidget(),   # frame chứa chart_temp
                self.ui.card_pressure_1.parentWidget(),    # frame chứa chart_pressure_a
                self.ui.card_pressure_2.parentWidget(),    # frame chứa chart_pressure_b
                self.ui.card_pressure_3.parentWidget(),    # frame chứa chart_pressure_c
            ]

        charts = [self.chart_temp, self.chart_pressure_a,
                self.chart_pressure_b, self.chart_pressure_c]

        for idx, chart in enumerate(charts):
            chart.clicked.connect(lambda i=idx: self._toggle_chart_frame(i))

        self._maximized_chart_idx = -1
        QTimer.singleShot(100, self._save_grid_rects)
        self._chart_layouts = [
            self.ui.card_temperature,
            self.ui.card_pressure_1,
            self.ui.card_pressure_2,
            self.ui.card_pressure_3,
        ]
        # QTimer.singleShot(25,  self.chart_pressure_a._render_timer.start)
        # QTimer.singleShot(50,  self.chart_pressure_b._render_timer.start)
        # QTimer.singleShot(75,  self.chart_pressure_c._render_timer.start)
        
    def _save_grid_rects(self):
        rects = [f.geometry() for f in self._chart_frames]
        if all(r.width() > 0 and r.height() > 0 for r in rects):
            self._grid_rects = rects
            # print("Grid rects saved:", rects)  # xoá sau khi debug xong
        else:
            QTimer.singleShot(100, self._save_grid_rects)
    
    def _toggle_chart_frame(self, idx: int):
        frames = self._chart_frames
        charts = [self.chart_temp, self.chart_pressure_a,
                self.chart_pressure_b, self.chart_pressure_c]
        grid = self.ui.gridLayout_2

        # idx → (row, col) trong grid 2x2
        positions = {0: (0, 0), 1: (0, 1), 2: (1, 0), 3: (1, 1)}
        row, col = positions[idx]

        if self._maximized_chart_idx == idx:
            # ── Thu nhỏ ───────────────────────────────────────────────────────
            self._maximized_chart_idx = -1

            # Trả stretch về đều
            grid.setRowStretch(0, 1)
            grid.setRowStretch(1, 1)
            grid.setColumnStretch(0, 1)
            grid.setColumnStretch(1, 1)

            # Hiện lại 3 frame còn lại
            for i, f in enumerate(frames):
                if i != idx:
                    f.show()

            # Bật lại render timer cho tất cả
            for c in charts:
                c._render_timer.start()

        else:
            # ── Phóng to ──────────────────────────────────────────────────────
            self._maximized_chart_idx = idx

            other_row = 1 if row == 0 else 0
            other_col = 1 if col == 0 else 0

            grid.setRowStretch(row,       100)
            grid.setRowStretch(other_row, 0)
            grid.setColumnStretch(col,       100)
            grid.setColumnStretch(other_col, 0)

            # Ẩn 3 frame còn lại + tắt render timer
            for i, f in enumerate(frames):
                if i != idx:
                    f.hide()
                    charts[i]._render_timer.stop()
        
    ###########################################################################################
    #############################------ Button Function Setup ------###########################
    def _setup_btn_signals(self):
        self.ui.left_side_menu_widget.customizeQCustomSlideMenu(
            defaultWidth = 70,
            defaultHeight = "parent",
            collapsedWidth = 70,
            collapsedHeight = "parent",
            expandedWidth = 175,
            expandedHeight = "parent",
            animationDuration = 200,
            animationEasingCurve = QEasingCurve.Linear,
            collapsingAnimationDuration = 200,
            collapsingAnimationEasingCurve = QEasingCurve.Linear,
            expandingAnimationDuration = 200,
            expandingAnimationEasingCurve = QEasingCurve.Linear,
        )
        self.ui.open_side_menu_btn.clicked.connect(self._on_menu_toggle)
        ### Left side menu button
        self.ui.home_page_btn.clicked.connect(self.pressure_page_btn)
        self.ui.chart_page_btn.clicked.connect(self.home_page_btn)
        self.ui.device_page_btn.clicked.connect(self.device_page_btn)
        self.ui.history_page_btn.clicked.connect(self.history_page_btn)

        self.ui.eng_language.clicked.connect(self.set_language_en)
        self.ui.cn_language.clicked.connect(self.set_language_cn)

        self.ui.next_group_page_btn.clicked.connect(self.next_previous_pressure_page)

        self.ui.back_connection_page_btn.clicked.connect(lambda: self.ui.stackedWidget_3.setCurrentWidget(self.ui.connection_page))
        self.ui.back_home_page_btn.clicked.connect(lambda: self.ui.home_page_btn.click())

        self.ui.temp_unit_selection_combox.currentIndexChanged.connect(lambda: QTimer.singleShot(50, self._set_cur_unit))

        self.ui.plc_io_btn.installEventFilter(self)
        self.ui.code_display.installEventFilter(self)

        self.ui.new_data_btn.clicked.connect(self.new_data_btn)
        
        self.ui.start_btn.clicked.connect(lambda: self.start_stop_btn(self.ui.start_btn))
        self.ui.stop_btn.clicked.connect(lambda: self.start_stop_btn(self.ui.stop_btn))

        # self.ui.heat_btn_a.toggled.connect(lambda checked: self.on_heat_btn_clicked("A", checked))
        # self.ui.heat_btn_b.toggled.connect(lambda checked: self.on_heat_btn_clicked("B", checked))
        # self.ui.heat_btn_c.toggled.connect(lambda checked: self.on_heat_btn_clicked("C", checked))

        self.ui.heat_btn_a.toggled.connect(lambda checked: self.heating_btn("A", checked, self.ui.heat_btn_a))
        self.ui.heat_btn_b.toggled.connect(lambda checked: self.heating_btn("B", checked, self.ui.heat_btn_b))
        self.ui.heat_btn_c.toggled.connect(lambda checked: self.heating_btn("C", checked, self.ui.heat_btn_c))
        self.ui.heat_btn_t0.toggled.connect(lambda checked: self.heating_btn("T0", checked, self.ui.heat_btn_t0))

        self.ui.vacuum_btn_a.toggled.connect(lambda checked: self.pumping_btn("A", checked, self.ui.vacuum_btn_a))
        self.ui.vacuum_btn_b.toggled.connect(lambda checked: self.pumping_btn("B", checked, self.ui.vacuum_btn_b))
        self.ui.vacuum_btn_c.toggled.connect(lambda checked: self.pumping_btn("C", checked, self.ui.vacuum_btn_c))

        self.ui.refuel_btn_a.toggled.connect(lambda checked: self.fill_oil_btn("A", checked, self.ui.refuel_btn_a))
        self.ui.refuel_btn_b.toggled.connect(lambda checked: self.fill_oil_btn("B", checked, self.ui.refuel_btn_b))
        self.ui.refuel_btn_c.toggled.connect(lambda checked: self.fill_oil_btn("C", checked, self.ui.refuel_btn_c))

        self.ui.set_cycle_a_btn.toggled.connect(lambda checked: self.cycle_loop_btn("A", checked, self.ui.set_cycle_a_btn))
        self.ui.set_cycle_b_btn.toggled.connect(lambda checked: self.cycle_loop_btn("B", checked, self.ui.set_cycle_b_btn))
        self.ui.set_cycle_c_btn.toggled.connect(lambda checked: self.cycle_loop_btn("C", checked, self.ui.set_cycle_c_btn))

        self.ui.clear_data_btn.clicked.connect(self.clear_data_btn)

        self.ui.reset_cycle_a_btn.clicked.connect(lambda: self.plc_writer_worker.write_value.emit("P1_Number_Test_Times", 0) if self.plc_writer_connection else None) #type: ignore
        self.ui.reset_cycle_b_btn.clicked.connect(lambda: self.plc_writer_worker.write_value.emit("P2_Number_Test_Times", 0) if self.plc_writer_connection else None) #type: ignore
        self.ui.reset_cycle_c_btn.clicked.connect(lambda: self.plc_writer_worker.write_value.emit("P3_Number_Test_Times", 0) if self.plc_writer_connection else None) #type: ignore

        self.ui.search_data.textChanged.connect(self._on_search_changed)
        self.ui.search_data_start_edit.dateTimeChanged.connect(self._on_search_changed)
        self.ui.search_data_end_edit.dateTimeChanged.connect(self._on_search_changed)
        self.ui.search_type_btn.clicked.connect(self.search_type_btn)
        self.ui.clear_history_search.clicked.connect(self._on_clear_clicked)
        self.ui.export_all_tables_to_excel_btn.clicked.connect(self.export_all_tables_to_excel_btn)

        spinbox_map = {
            "pressure_sv_a_1": self.on_pressure_sv_a_1_changed,
            "pressure_sv_a_5": self.on_pressure_sv_a_5_changed,
            "pressure_sv_a_6": self.on_pressure_sv_a_6_changed,
            "pressure_sv_a_7": self.on_pressure_sv_a_7_changed,
            "pressure_sv_a_8": self.on_pressure_sv_a_8_changed,
            "pressure_sv_a_9": self.on_pressure_sv_a_9_changed,
            "pressure_sv_a_10": self.on_pressure_sv_a_10_changed,
            "pressure_sv_a_11": self.on_cycle_a_displ_2_changed,

            "pressure_sv_b_1": self.on_pressure_sv_b_1_changed,
            "pressure_sv_b_5": self.on_pressure_sv_b_5_changed,
            "pressure_sv_b_6": self.on_pressure_sv_b_6_changed,
            "pressure_sv_b_7": self.on_pressure_sv_b_7_changed,
            "pressure_sv_b_8": self.on_pressure_sv_b_8_changed,
            "pressure_sv_b_9": self.on_pressure_sv_b_9_changed,
            "pressure_sv_b_10": self.on_pressure_sv_b_10_changed,
            "pressure_sv_b_11": self.on_cycle_b_displ_2_changed,

            "pressure_sv_c_1": self.on_pressure_sv_c_1_changed,
            "pressure_sv_c_5": self.on_pressure_sv_c_5_changed,
            "pressure_sv_c_6": self.on_pressure_sv_c_6_changed,
            "pressure_sv_c_7": self.on_pressure_sv_c_7_changed,
            "pressure_sv_c_8": self.on_pressure_sv_c_8_changed,
            "pressure_sv_c_9": self.on_pressure_sv_c_9_changed,
            "pressure_sv_c_10": self.on_pressure_sv_c_10_changed,
            "pressure_sv_c_11": self.on_cycle_c_displ_2_changed,

            "t0_sv": self.on_t0_sv_changed,
            "at_sv": self.on_at_sv_changed,
            "bt_sv": self.on_bt_sv_changed,
            "ct_sv": self.on_ct_sv_changed,

            "t0_h_alm_value": self.on_t0_h_alm_value_changed,
            "at_h_alm_value": self.on_at_h_alm_value_changed,
            "bt_h_alm_value": self.on_bt_h_alm_value_changed,
            "ct_h_alm_value": self.on_ct_h_alm_value_changed,

            "t0_l_alm_value": self.on_t0_l_alm_value_changed,
            "at_l_alm_value": self.on_at_l_alm_value_changed,
            "bt_l_alm_value": self.on_bt_l_alm_value_changed,
            "ct_l_alm_value": self.on_ct_l_alm_value_changed,

            "t0_offset_value": self.on_t0_offset_value_changed,
            "at_t1_offset_value": self.on_at_t1_offset_value_changed,
            "at_t2_offset_value": self.on_at_t2_offset_value_changed,
            "at_t3_offset_value": self.on_at_t3_offset_value_changed,
            "bt_t1_offset_value": self.on_bt_t1_offset_value_changed,
            "bt_t2_offset_value": self.on_bt_t2_offset_value_changed,
            "bt_t3_offset_value": self.on_bt_t3_offset_value_changed,
            "ct_t1_offset_value": self.on_ct_t1_offset_value_changed,
            "ct_t2_offset_value": self.on_ct_t2_offset_value_changed,
            "ct_t3_offset_value": self.on_ct_t3_offset_value_changed,

            "table_write_cycle": self.on_table_write_cycle_value_changed,

        }

        install_clear_on_focus(self.ui.db_file_path_edit)
        install_clear_on_focus(self.ui.plc_ip_address_edit)
        install_clear_on_focus(self.ui.db_number_input)
        install_clear_on_focus(self.ui.db_data_size_input)

        protect_widgets_on_stacked(self.ui.stackedWidget_2, [
            (self.ui.db_file_path_edit,   None, None),
            (self.ui.plc_ip_address_edit, None, None),
            (self.ui.db_number_input,     None, None),
            (self.ui.db_data_size_input,  None, None)
        ])

        for name, handler in spinbox_map.items():
            spinbox = getattr(self.ui, name)
            install_clear_on_focus(spinbox)
            spinbox.setKeyboardTracking(False)
            spinbox.valueChanged.connect(handler)

        self.ui.error_display._speed = 50

    def _init_history_database(self):
        """Khởi tạo SQLite Database - Không xóa bảng cũ"""
        db_folder = Path(self.stk_mch_folder) / "DB_address"
        db_folder.mkdir(parents=True, exist_ok=True)
        
        self.history_db_path = db_folder / "history.db"
        
        try:
            self.conn = sqlite3.connect(str(self.history_db_path), check_same_thread=False)
            self.conn.execute("PRAGMA journal_mode=WAL")
            self.conn.execute("PRAGMA synchronous=NORMAL")
            # self.conn.execute("DROP TABLE IF EXISTS history")
            # Kiểm tra xem bảng đã có chưa
            cursor = self.conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='history'")
            table_exists = cursor.fetchone() is not None

            if not table_exists:
                # Tạo mới nếu chưa có
                self.conn.execute('''
                    CREATE TABLE history (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        "No." TEXT,
                        "Name." TEXT,
                        "Group." TEXT,
                        "Pressure." TEXT,
                        "T-Oven." TEXT,
                        "Front." TEXT,
                        "Middle." TEXT,
                        "End." TEXT,
                        "Date." TEXT
                    )
                ''')
                # self.logger.info("Created new history table")
            else:
                # self.logger.info("History table already exists")

                self.conn.commit()
                self.logger.info(f"SQLite DB ready: {self.history_db_path}")
            
        except Exception as e:
            # self.logger.error(f"SQLite init error: {e}")
            self.conn = None

    def _init_table_list_history(self):
        if not self.conn:
            return
        try:
            cursor = self.conn.cursor()
            cursor.execute('SELECT * FROM history ORDER BY id ASC')  # ✅ ASC
            rows = cursor.fetchall()

            # Cache toàn bộ theo đúng thứ tự ASC
            self._all_rows_cache = []
            for idx, row in enumerate(rows):
                formatted = [str(idx + 1)] + [str(v if v is not None else "") for v in row[2:]]
                self._all_rows_cache.append(formatted)

            total = len(self._all_rows_cache)
            self._displayed_start = max(0, total - self._table_display)
            
            self._render_chunk(self._displayed_start, total)

            self.ui.list_history.verticalScrollBar().valueChanged.connect(
                self._on_history_scroll
            )
            self._resize_table_columns(self.ui.list_history)
            self.ui.list_history.scrollToBottom()
            self.logger.info(f"Cached {total:,} records, showing last {self._table_display}")

        except Exception as e:
            self.logger.error(f"Error loading history: {e}")

    def _render_chunk(self, start: int, end: int):
        """Render cache[start:end] vào table"""
        chunk = self._all_rows_cache[start:end]
        if not chunk:
            return

        table = self.ui.list_history
        table.setUpdatesEnabled(False)
        for row_data in chunk:
            row_pos = table.rowCount()
            table.insertRow(row_pos)
            for col, value in enumerate(row_data):
                item = QTableWidgetItem(value)
                if col == 0:
                    item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row_pos, col, item)
        table.setUpdatesEnabled(True)

    def _on_history_scroll(self, value: int):
        if value <= 100:
            if self._displayed_start <= 0:
                return
            if getattr(self, '_loading_history_chunk', False):
                return
            self._loading_history_chunk = True
            QTimer.singleShot(1000, self._prepend_history_chunk)

        self._reset_scroll_timer()

    def _prepend_history_chunk(self):
        try:
            if self._displayed_start <= 0:
                return

            table = self.ui.list_history
            sb = table.verticalScrollBar()

            # Nếu user đã scroll xuống rồi thì thôi
            if sb.value() > 10:
                return

            sb.valueChanged.disconnect(self._on_history_scroll)

            new_start = max(0, self._displayed_start - self._table_display)
            chunk = self._all_rows_cache[new_start:self._displayed_start]

            table.setUpdatesEnabled(False)
            for i, row_data in enumerate(chunk):
                table.insertRow(i)
                for col, value in enumerate(row_data):
                    item = QTableWidgetItem(value)
                    if col == 0:
                        item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(i, col, item)
            table.setUpdatesEnabled(True)

            self._displayed_start = new_start

            first_visible_row = len(chunk)
            table.scrollTo(
                table.model().index(first_visible_row, 0),
                QTableWidget.ScrollHint.PositionAtTop
            )

            sb.valueChanged.connect(self._on_history_scroll)

        finally:
            self._loading_history_chunk = False

    def _reset_scroll_timer(self):
        """Mỗi lần scroll đều reset timer — sau 10s không scroll thì về 100 dòng cuối"""
        if not hasattr(self, '_scroll_reset_timer'):
            self._scroll_reset_timer = QTimer(self)
            self._scroll_reset_timer.setSingleShot(True)
            self._scroll_reset_timer.timeout.connect(self._reset_to_latest_100)

        self._scroll_reset_timer.stop()
        self._scroll_reset_timer.start(10000)  # 10s

    def _reset_to_latest_100(self):
        """Reset về 100 dòng cuối nếu không tương tác 10s"""
        table = self.ui.list_history
        sb = table.verticalScrollBar()

        # Nếu đã ở bottom rồi thì không cần reset
        if sb.value() >= sb.maximum() - 5:
            return

        total = len(self._all_rows_cache)
        new_start = max(0, total - 100)

        if new_start == self._displayed_start and table.rowCount() <= 100:
            return  # Đã đúng rồi

        sb.valueChanged.disconnect(self._on_history_scroll)

        table.setUpdatesEnabled(False)
        table.setRowCount(0)

        self._displayed_start = new_start
        self._render_chunk(new_start, total)

        table.setUpdatesEnabled(True)
        table.scrollToBottom()

        sb.valueChanged.connect(self._on_history_scroll)

    def _on_search_changed(self, *args):
        """Hỗ trợ cả textbox và dateTimeChanged"""
        # Lấy keyword từ textbox
        keyword = self.ui.search_data.text().strip().lower()
        
        if not self.conn:
            return

        # Nếu không có từ khóa và không phải chế độ name-group-date → load mặc định
        start_dt = self.ui.search_data_start_edit.dateTime()
        end_dt = self.ui.search_data_end_edit.dateTime()
        if start_dt == end_dt and not keyword:
            self.ui.clear_history_search.hide()
            self.ui.stackedWidget_4.setCurrentIndex(0)
            return
            
        self.ui.clear_history_search.show()
        self.ui.stackedWidget_4.setCurrentIndex(1)

        try:
            where_clauses = []
            params = []

            # ==================== TỪ KHÓA ====================
            if keyword:
                if self._current_search_type == "name":
                    where_clauses.append('"Name." LIKE ?')
                    params.append(f"%{keyword}%")
                elif self._current_search_type == "group":
                    where_clauses.append('"Group." LIKE ?')
                    params.append(f"%{keyword}%")
                else:  # name-group-date
                    if keyword.startswith("-"):
                        group_part = keyword[1:].strip()
                        where_clauses.append('"Group." LIKE ?')
                        params.append(f"%{group_part}%")

                    # "name - group name" → tìm AND cả Name lẫn Group
                    elif " - " in keyword:
                        parts = keyword.split(" - ", 1)
                        name_part = parts[0].strip()
                        group_part = parts[1].strip()
                        where_clauses.append('"Name." LIKE ? AND "Group." LIKE ?')
                        params.extend([f"%{name_part}%", f"%{group_part}%"])

                    # Không có "-" → tìm OR như cũ
                    else:
                        where_clauses.append('("Name." LIKE ? OR "Group." LIKE ?)')
                        params.extend([f"%{keyword}%", f"%{keyword}%"])

            # ==================== KHOẢNG THỜI GIAN ====================
            if self._current_search_type == "name-group-date":
                start_date = self.ui.search_data_start_edit.dateTime().toString("HH:mm - dd/MM/yyyy")
                end_date = self.ui.search_data_end_edit.dateTime().toString("HH:mm - dd/MM/yyyy")

                if start_date and end_date:
                    where_clauses.append('"Date." >= ? AND "Date." <= ?')
                    params.extend([start_date, end_date])

            # Nếu không có điều kiện nào
            if not where_clauses:
                self.ui.clear_history_search.hide()
                self.ui.stackedWidget_4.setCurrentIndex(0)
                # self._init_table_list_history()
                return

            where_clause = " AND ".join(where_clauses)
            self._search_where_clause = where_clause
            self._search_params = tuple(params)

            self._search_fetch_chunk(reset=True)

        except Exception as e:
            self.logger.error(f"Search error: {e}")
            # self._init_table_list_history()

    def _search_fetch_chunk(self, reset=False):
        if not hasattr(self, '_search_where_clause'):
            return

        where_clause = self._search_where_clause
        params = self._search_params

        try:
            cursor = self.conn.cursor()
            
            # === QUAN TRỌNG: Nếu là search theo thời gian thì lấy HẾT ===
            if self._current_search_type == "name-group-date":
                # Lấy toàn bộ kết quả theo thời gian (không giới hạn)
                cursor.execute(
                    f'SELECT * FROM history WHERE {where_clause} ORDER BY id ASC',
                    params
                )
            else:
                # Các chế độ khác thì giới hạn 200 để tránh lag
                cursor.execute(
                    f'SELECT * FROM history WHERE {where_clause} ORDER BY id ASC',  # LIMIT 200
                    params
                )

            rows = cursor.fetchall()

            filtered = []
            for row_idx, row in enumerate(rows):
                # row[0] = id, row[1] = No.
                if row[1] is not None and str(row[1]).strip() != "":
                    original_stt = str(row[1])
                else:
                    original_stt = str(row[0])  # Fallback về id

                row_list = [original_stt] + [str(v if v is not None else "") for v in row[2:]]
                filtered.append(row_list)

            self._load_filtered_data(filtered)
        except Exception as e:
            self.logger.error(f"_search_fetch_chunk error: {e}")
            # self._init_table_list_history()

    def _on_search_scroll(self, value: int):
        """
        Xử lí dữ liệu khi cuộn History lên trên, thêm 100 hoặc tuỳ chỉnh số dòng vào table History
        """
        if value > 100:
            return
        if self._search_db_offset <= 0:
            return
        if getattr(self, '_loading_search_chunk', False):
            return

        self._loading_search_chunk = True
        QTimer.singleShot(2000, self._search_fetch_chunk)

    def _load_filtered_data(self, filtered_data: list):
        """
        Load dữ liệu từ database vào table thứ 2 để hiển thị
        """
        table = self.ui.list_history_2
        table.setUpdatesEnabled(False)
        table.setRowCount(0)
        table.setRowCount(len(filtered_data))

        for row_idx, row_data in enumerate(filtered_data):
            for col, value in enumerate(row_data):
                item = QTableWidgetItem(value)
                if col == 0:
                    item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row_idx, col, item)

        table.setUpdatesEnabled(True)
        self._resize_table_columns(table)
        # self._update_info(self._search_total, len(self._all_rows_cache) + self._db_offset)

    def _resize_table_columns(self, table):
        if not table or table.columnCount() == 0:
            return
            
        header = table.horizontalHeader()
        for col in range(table.columnCount() - 1):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.ResizeToContents)

    def _update_info(self, shown: int, total: int):
        if shown == total:
            self.ui.label_info.setText(f"")
        else:
            self.ui.label_info.setText(f"Found {shown} / {total} records")

    def search_type_btn(self):
        if self._current_search_type == "name":
            self._current_search_type = "group"
            self.ui.search_data.setPlaceholderText("Search [Group]")
        elif self._current_search_type == "group":
            self._current_search_type = "name-group-date"
            self.ui.search_data.setSizePolicy(
                QSizePolicy.Policy.Preferred,
                QSizePolicy.Policy.Preferred
            )
            self.ui.search_data.setPlaceholderText("[Name]-[Group]")
            self.ui.multi_search_data.show()
        elif self._current_search_type == "name-group-date":
            self._current_search_type = "name"
            self._set_time_search_data_start_edit()
            self._set_time_search_data_end_edit()
            self.ui.search_data.setSizePolicy(
                QSizePolicy.Policy.Expanding,
                QSizePolicy.Policy.Preferred
            )
            self.ui.search_data.setPlaceholderText("Search [Name]")
            self.ui.multi_search_data.hide()

    def _on_clear_clicked(self):
        self.ui.search_data.clear()
        self._set_time_search_data_start_edit()
        self._set_time_search_data_end_edit()
        self.ui.search_data.setFocus()
        self.ui.list_history.scrollToBottom()

    def _set_time_search_data_start_edit(self):
        now = QDateTime.currentDateTime()
        rounded_time, extra_day = self.round_time(now.time())
        
        self.ui.search_data_start_edit.setDateTime(
            QDateTime(now.date().addDays(extra_day), rounded_time)
        )

    def _set_time_search_data_end_edit(self):
        now = QDateTime.currentDateTime()
        rounded_time, extra_day = self.round_time(now.time())
        
        self.ui.search_data_end_edit.setDateTime(
            QDateTime(now.date().addDays(extra_day), rounded_time)
        )

    def round_time(self, time: QTime):
        minute = time.minute()
        hour = time.hour()

        if minute <= 30:
            return QTime(hour, 30, 0), 0        # (time, ngày cộng thêm)
        else:
            next_hour = hour + 1
            if next_hour >= 24:
                return QTime(0, 0, 0), 1        # Qua ngày mới
            return QTime(next_hour, 0, 0), 0
        
    # ── Group A ──────────────────────────────────────────
    def on_pressure_sv_a_1_changed(self, value: float): 
        if self.plc_writer_connection:
            self.plc_writer_worker.write_value.emit("P1_TemperatureSetting", self.cal_fah_to_cel(value)) #type: ignore
        else: 
            pass
        self.ui.at_sv.blockSignals(True)
        self.ui.at_sv.setValue(value)
        self.ui.at_sv.blockSignals(False)

    def on_pressure_sv_a_5_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_PressureSetting", value) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_a_6_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_Air_FillingTime", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_a_7_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_Air_HoldingTime", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_a_8_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_Air_ReleaseTime", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_a_9_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_Oil_Start_Time", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_a_10_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_Oil_End_Time", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore

    # ── Group B ──────────────────────────────────────────
    def on_pressure_sv_b_1_changed(self, value: float): 
        if self.plc_writer_connection:
            self.plc_writer_worker.write_value.emit("P2_TemperatureSetting", self.cal_fah_to_cel(value)) #type: ignore
        else: 
            pass
        self.ui.bt_sv.blockSignals(True)
        self.ui.bt_sv.setValue(value)
        self.ui.bt_sv.blockSignals(False)

    def on_pressure_sv_b_5_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_PressureSetting", value) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_b_6_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_Air_FillingTime", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_b_7_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_Air_HoldingTime", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_b_8_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_Air_ReleaseTime", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_b_9_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_Oil_Start_Time", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_b_10_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_Oil_End_Time", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore

    # ── Group C ──────────────────────────────────────────
    def on_pressure_sv_c_1_changed(self, value: float): 
        if self.plc_writer_connection:
            self.plc_writer_worker.write_value.emit("P3_TemperatureSetting", self.cal_fah_to_cel(value)) #type: ignore
        else: 
            pass
        self.ui.ct_sv.blockSignals(True)
        self.ui.ct_sv.setValue(value)
        self.ui.ct_sv.blockSignals(False)

    def on_pressure_sv_c_5_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_PressureSetting", value) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_c_6_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_Air_FillingTime", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_c_7_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_Air_HoldingTime", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_c_8_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_Air_ReleaseTime", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_c_9_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_Oil_Start_Time", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore
    def on_pressure_sv_c_10_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_Oil_End_Time", self.cal_sec_to_msec(value)) if self.plc_writer_connection else None #type: ignore

    # ── Cycle ──────────────────────────────────────────
    def on_cycle_a_displ_2_changed(self, value: int): self.plc_writer_worker.write_value.emit("P1_CountTimes", value) if self.plc_writer_connection else None #type: ignore
    def on_cycle_b_displ_2_changed(self, value: int): self.plc_writer_worker.write_value.emit("P2_CountTimes", value) if self.plc_writer_connection else None #type: ignore
    def on_cycle_c_displ_2_changed(self, value: int): self.plc_writer_worker.write_value.emit("P3_CountTimes", value) if self.plc_writer_connection else None #type: ignore

    # ── Temperature Modify ──────────────────────────────────────────
    def on_t0_sv_changed(self, value: float): self.plc_writer_worker.write_value.emit("T0_TemperatureSetting", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_at_sv_changed(self, value: float): self.ui.pressure_sv_a_1.setValue(value)
    def on_bt_sv_changed(self, value: float): self.ui.pressure_sv_b_1.setValue(value)
    def on_ct_sv_changed(self, value: float): self.ui.pressure_sv_c_1.setValue(value)

    def on_t0_h_alm_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("T0_TempLimitHIGH", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_at_h_alm_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_TempLimitHIGH", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_bt_h_alm_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_TempLimitHIGH", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_ct_h_alm_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_TempLimitHIGH", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore

    def on_t0_l_alm_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("T0_TempLimitLOW", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_at_l_alm_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_TempLimitLOW", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_bt_l_alm_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_TempLimitLOW", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_ct_l_alm_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_TempLimitLOW", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore

    def on_t0_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("T0_TempOffset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_at_t1_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_Temp1Offset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_at_t2_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_Temp2Offset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_at_t3_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P1_Temp3Offset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_bt_t1_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_Temp1Offset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_bt_t2_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_Temp2Offset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_bt_t3_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P2_Temp3Offset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_ct_t1_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_Temp1Offset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_ct_t2_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_Temp2Offset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_ct_t3_offset_value_changed(self, value: float): self.plc_writer_worker.write_value.emit("P3_Temp3Offset", self.cal_fah_to_cel(value)) if self.plc_writer_connection else None #type: ignore
    def on_table_write_cycle_value_changed(self, value: float): self.ui.table_write_cycle.setValue(float(self.ui.read_time_input.value())/1000) if (value*1000) < self.ui.read_time_input.value() else None #type: ignore

    def _on_menu_toggle(self):
        # Pause tất cả chart
        if self.ui.stackedWidget_2.currentIndex() == 0:
            self._pause_charts()
            self.ui.left_side_menu_widget.slideMenu()
            
            # Resume sau khi animation xong
            QTimer.singleShot(220, self._resume_charts)
        else:
            self.ui.left_side_menu_widget.slideMenu()

    def _pause_charts(self): 
        self.chart_temp._render_timer.stop()
        self.chart_pressure_a._render_timer.stop()
        self.chart_pressure_b._render_timer.stop()
        self.chart_pressure_c._render_timer.stop()

    def _resume_charts(self):
        self.chart_temp._render_timer.start()
        self.chart_pressure_a._render_timer.start()
        self.chart_pressure_b._render_timer.start()
        self.chart_pressure_c._render_timer.start()

    def home_page_btn(self):
        self.user = False
        self.ui.stackedWidget_2.setCurrentWidget(self.ui.menu_page)

    def pressure_page_btn(self):
        self.user = False
        self.ui.stackedWidget_2.setCurrentWidget(self.ui.temp_press_page)
        self.ui.stackedWidget.setCurrentWidget(self.ui.pressure_page)

    def temperature_page_btn(self):
        self.user = False
        self.ui.home_page_btn.click()
        self.ui.stackedWidget.setCurrentWidget(self.ui.temperature_page)
        
    def device_page_btn(self):
        # self.user = False
        self.ui.stackedWidget_2.setCurrentWidget(self.ui.device_page)
        self.connection_page_btn()

    def connection_page_btn(self):
        self.ui.stackedWidget_3.setCurrentWidget(self.ui.connection_page)
    
    def i_o_page_btn(self):
        self.ui.stackedWidget_3.setCurrentWidget(self.ui.i_o_page)
    
    def history_page_btn(self):
        self.user = False
        self.ui.stackedWidget_2.setCurrentWidget(self.ui.history_page)

    # def (self):
    
    def next_previous_pressure_page(self):
        index = self.ui.stackedWidget.currentIndex()
        if index == 0:
            self.ui.stackedWidget.setCurrentIndex(1)
        elif index == 1:
            self.ui.stackedWidget.setCurrentIndex(0)

    def _setup_plc_threads(self, state: bool):
        if state:
            self.logger.info("PLC Thread set Off")
            self.setup_simulate_threads()
            return
        
        if not self.db_dict:
            ltmessage.error(self, "Error", "DB Layout not found! Cannot start PLC threads.")
            return 

        if not self._setup_write_plc_thread(
            ip=self.db_dict["ip_plc"],
            db_number=self.db_dict["db_name"],
            db_layout=self.db_dict["DB_LAYOUT"],
            db_size=self.db_dict["DB_TOTAL_BYTES"],
            poll_ms=self.db_dict["write_time"]
        ):
            ltmessage.error(self, "Error", "Failed to connect to PLC! Try again later.")

        time.sleep(0.2)

        if not self._setup_read_plc_thread(
            ip=self.db_dict["ip_plc"],
            db_number=self.db_dict["db_name"],
            db_layout=self.db_dict["DB_LAYOUT"],
            db_size=self.db_dict["DB_TOTAL_BYTES"],
            poll_ms=self.db_dict["read_time"]
        ):
            ltmessage.error(self, "Error", "Failed to connect to PLC! Try again later.")

    def setup_simulate_threads(self):
        try:
            simulate_thread = DataSimulator()
            simulate_thread.db_data_convert.connect(self._data_group_filter)
            simulate_thread.start()
            self.thread_dict["data_simulator"] = simulate_thread

            if not simulate_thread.isRunning():
                raise Exception("DataSimulator thread failed to start")

            time.sleep(0.1)

        except Exception as e:
            return 

    def _setup_read_plc_thread(
            self, 
            ip: str = "172.16.100.100", 
            db_number: Optional[int] = None, 
            db_layout: Optional[list[tuple[str, str, int, Any]]] = None, 
            db_size: Optional[int] = None, 
            poll_ms: int = 500
            ):
        try:
            if db_number is None:
                raise ValueError("DB number is not defined. Cannot start PLC read thread.")
            elif db_layout is None:
                raise ValueError("DB layout is not defined. Cannot start PLC read thread.")
            elif db_size is None:
                raise ValueError("DB size is not defined. Cannot start PLC read thread.")
            
            self.plc_read_thread = QThread()
            self.plc_read_worker = PLCRead(
                ip=ip,
                db_number=db_number,
                db_layout=db_layout,
                db_size=db_size,
                poll_ms=poll_ms
            )
            self.plc_read_worker.moveToThread(self.plc_read_thread)

            self.plc_read_thread.started.connect(self.plc_read_worker.run)
            self.plc_read_worker.data_ready.connect(self._data_ready)
            self.plc_read_worker.connected.connect(self._read_status_plc)
            self.plc_read_worker.init_data.connect(self._set_system_data)
            
            self.plc_read_worker.finished.connect(self.plc_read_thread.quit)
            self.plc_read_worker.finished.connect(self.plc_read_worker.deleteLater)
            self.plc_read_thread.finished.connect(self.plc_read_thread.deleteLater)

            self.plc_read_thread.start()
            if not self.plc_read_thread.isRunning():
                raise Exception("plc_read_thread failed to start")
        except Exception as e:
            self.logger.info("PLC Reader gone wrong:", e)
            return False
        
        self.worker_dict["plc_read_worker"] = self.plc_read_worker
        return True

    def _setup_write_plc_thread(
            self, 
            ip: str = "172.100.16.100", 
            db_number: Optional[int] = None, 
            db_layout: Optional[List] = None, 
            db_size: Optional[int] = None, 
            poll_ms: int = 500
            ):
        try:
            if db_number is None:
                raise ValueError("DB number is not defined. Cannot start PLC write thread.")
            elif db_layout is None:
                raise ValueError("DB layout is not defined. Cannot start PLC write thread.")
            elif db_size is None:
                raise ValueError("DB size is not defined. Cannot start PLC write thread.")

            self.plc_writer_thread = QThread()
            self.plc_writer_worker = PLCWrite(
                ip=ip, 
                db_number=db_number, 
                db_layout=db_layout,
                db_size=db_size,
                write_gap_ms=poll_ms
            )
            self.plc_writer_worker.moveToThread(self.plc_writer_thread)
            self.plc_writer_thread.started.connect(self.plc_writer_worker.run)
            self.plc_writer_worker.connected.connect(self._write_status_plc)

            self.plc_writer_worker.finished.connect(self.plc_writer_thread.quit)
            self.plc_writer_worker.finished.connect(self.plc_writer_worker.deleteLater)
            self.plc_writer_thread.finished.connect(self.plc_writer_thread.deleteLater)

            self.plc_writer_thread.start()
            
        except Exception as e:
            self.logger.info("PLC Writer gone wrong:", e)
        self.thread_dict["plc_writer_thread"] = self.plc_writer_thread
        return True
    
    def _data_ready(self, data: dict):
        # def _t(label, fn):
            # t = time.perf_counter()
            # fn()
            # ms = (time.perf_counter() - t) * 1000
            # if ms > 1:
                # print(f"  [{label}] {ms:.1f}ms")

        try:
            if self.init_signal:
                self.logger.info("[Main]-[_data_ready]: Getting PLC State")
                self._init_pressure_group_sv_obj([
                    int(data.get('P1_CountTimes', 0)),
                    float(data.get('P1_Oil_Start_Time', 0)/1000),
                    float(data.get('P1_Oil_End_Time', 0)/1000),
                    float(data.get('P1_Air_FillingTime', 0)/1000),
                    float(data.get('P1_Air_HoldingTime', 0)/1000),
                    float(data.get('P1_Air_ReleaseTime', 0)/1000),
                    float(data.get('P1_PressureSetting', 0.00)),
                    self.for_display_temp(float(data.get('P1_TemperatureSetting', 0.0))),
                    self.for_display_temp(float(data.get('P1_TempLimitHIGH', 0.0))),
                    self.for_display_temp(float(data.get('P1_TempLimitLOW', 0.0))),
                    self.for_display_temp(float(data.get('P1_Temp1Offset', 0.0))),
                    self.for_display_temp(float(data.get('P1_Temp2Offset', 0.0))),
                    self.for_display_temp(float(data.get('P1_Temp3Offset', 0.0)))
                ],
                [
                    int(data.get('P2_CountTimes', 0)),
                    float(data.get('P2_Oil_Start_Time', 0)/1000),
                    float(data.get('P2_Oil_End_Time', 0)/1000),
                    float(data.get('P2_Air_FillingTime', 0)/1000),
                    float(data.get('P2_Air_HoldingTime', 0)/1000),
                    float(data.get('P2_Air_ReleaseTime', 0)/1000),
                    float(data.get('P2_PressureSetting', 0.00)),
                    float(data.get('P2_TemperatureSetting', 0.0)),
                    float(data.get('P2_TempLimitHIGH', 0.0)),
                    float(data.get('P2_TempLimitLOW', 0.0)),
                    float(data.get('P2_Temp1Offset', 0.0)),
                    float(data.get('P2_Temp2Offset', 0.0)),
                    float(data.get('P2_Temp3Offset', 0.0))
                ],
                [
                    int(data.get('P3_CountTimes', 0)),
                    float(data.get('P3_Oil_Start_Time', 0)/1000),
                    float(data.get('P3_Oil_End_Time', 0)/1000),
                    float(data.get('P3_Air_FillingTime', 0)/1000),
                    float(data.get('P3_Air_HoldingTime', 0)/1000),
                    float(data.get('P3_Air_ReleaseTime', 0)/1000),
                    float(data.get('P3_PressureSetting', 0.00)),
                    float(data.get('P3_TemperatureSetting', 0.0)),
                    self.for_display_temp(float(data.get('P3_TemperatureSetting', 0.0))),
                    self.for_display_temp(float(data.get('P3_TempLimitHIGH', 0.0))),
                    self.for_display_temp(float(data.get('P3_TempLimitLOW', 0.0))),
                    self.for_display_temp(float(data.get('P3_Temp1Offset', 0.0))),
                    self.for_display_temp(float(data.get('P3_Temp2Offset', 0.0))),
                    self.for_display_temp(float(data.get('P3_Temp3Offset', 0.0)))
                ],
                [
                    self.for_display_temp(float(data.get('T0_TemperatureSetting', 0.0))),
                    self.for_display_temp(float(data.get('T0_TempLimitHIGH', 0.0))),
                    self.for_display_temp(float(data.get('T0_TempLimitLOW', 0.0))),
                    self.for_display_temp(float(data.get('T0_TempOffset', 0.0)))
                ])
                print("Init Value Done")
                
                self._init_button_obj([
                    bool(data.get('START', False)),
                    bool(data.get('STOP', False)),
                    bool(data.get('T0_Start_Heat', False)),
                    bool(data.get('T0_Stop_Heat', False)),
                    bool(data.get('P1_Start_Heat', False)),
                    bool(data.get('P1_Start_Pressure', False)),
                    bool(data.get('P1_Start_Oil', False)),
                    bool(data.get('P1_BitCountTimes', False)),
                    bool(data.get('P2_Start_Heat', False)),
                    bool(data.get('P2_Start_Pressure', False)),
                    bool(data.get('P2_Start_Oil', False)),
                    bool(data.get('P2_BitCountTimes', False)),
                    bool(data.get('P3_Start_Heat', False)),
                    bool(data.get('P3_Start_Pressure', False)),
                    bool(data.get('P3_Start_Oil', False)),
                    bool(data.get('P3_BitCountTimes', False)),
                ])
                print("Init Button Done")
                self.init_signal = False

            now = time.time()
            if now - self._last_history_time >= self.ui.table_write_cycle.value():
                self._last_history_time = now
                if bool(data.get('P1_Start_Heat', False)) or bool(data.get('P1_Start_Pressure', False)):
                    self.add_row_to_list_history(
                        self.ui.code_display.text(),
                        "Group A",
                        float(data.get('P1_Current_PressureHose', 0.0)),
                        float(data.get('T0_Current_Temp', 0.0)),
                        float(data.get('P1_Current_Temp1', 0.0)), 
                        float(data.get('P1_Current_Temp2', 0.0)), 
                        float(data.get('P1_Current_Temp3', 0.0))
                    )
                if bool(data.get('P2_Start_Heat', False)) or bool(data.get('P2_Start_Pressure', False)):
                    self.add_row_to_list_history(
                        self.ui.code_display.text(),
                        "Group B",
                        float(data.get('P2_Current_PressureHose', 0.00)),
                        float(data.get('T0_Current_Temp', 0.0)),
                        float(data.get('P2_Current_Temp1', 0.0)), 
                        float(data.get('P2_Current_Temp2', 0.0)), 
                        float(data.get('P2_Current_Temp3', 0.0))
                    )
                if bool(data.get('P3_Start_Heat', False)) or bool(data.get('P3_Start_Pressure', False)):
                    self.add_row_to_list_history(
                        self.ui.code_display.text(),
                        "Group C",
                        float(data.get('P3_Current_PressureHose', 0.0)),
                        float(data.get('T0_Current_Temp', 0.0)),
                        float(data.get('P3_Current_Temp1', 0.0)), 
                        float(data.get('P3_Current_Temp2', 0.0)), 
                        float(data.get('P3_Current_Temp3', 0.0))
                    )

            # _t("t0_input_heat", lambda: 
            self._t0_input_heat_filter([
                bool(data.get('T0_Start_Heat', False)),
                bool(data.get('T0_Stop_Heat', False))
            ])
            # )

            # _t("input_data", lambda: 
            self._input_data_filter([
                bool(data.get('P1_Start_Heat', False)),
                bool(data.get('P1_Start_Pressure', False)),
                bool(data.get('P1_Start_Oil', False)),
                bool(data.get('P1_BitCountTimes', False)),
                bool(data.get('P2_Start_Heat', False)),
                bool(data.get('P2_Start_Pressure', False)),
                bool(data.get('P2_Start_Oil', False)),
                bool(data.get('P2_BitCountTimes', False)),
                bool(data.get('P3_Start_Heat', False)),
                bool(data.get('P3_Start_Pressure', False)),
                bool(data.get('P3_Start_Oil', False)),
                bool(data.get('P3_BitCountTimes', False)),
            ])
            # )

            # _t("i_o_group_3", lambda: 
            self._i_o_group_3_filter([
                float(data.get('T0_Current_Temp', 0.0)),
                float(data.get('P1_Current_Temp1', 0.0)),
                float(data.get('P1_Current_Temp2', 0.0)),
                float(data.get('P1_Current_Temp3', 0.0)),
                float(data.get('P2_Current_Temp1', 0.0)),
                float(data.get('P2_Current_Temp2', 0.0)),
                float(data.get('P2_Current_Temp3', 0.0)),
                float(data.get('P3_Current_Temp1', 0.0)),
                float(data.get('P3_Current_Temp2', 0.0)),
                float(data.get('P3_Current_Temp3', 0.0)),
                float(data.get('P1_Current_PressureHose', 0.00)),
                float(data.get('P2_Current_PressureHose', 0.00)),
                float(data.get('P3_Current_PressureHose', 0.00)),
                float(data.get('P1_Current_PressureITV', 0.0)),
                float(data.get('P2_Current_PressureITV', 0.0)),
                float(data.get('P3_Current_PressureITV', 0.0))
            ])
            # )

            # _t("t0_data", lambda: 
            self._t0_data_filter([
                float(data.get('T0_TemperatureSetting', 0.0)),
                float(data.get('T0_Current_Temp', 0.0))
            ])
            # )

            # _t("group_a", lambda: 
            self._group_a_data_filter([
                float(data.get('P1_Current_Temp1', 0.0)),
                float(data.get('P1_Current_Temp2', 0.0)),
                float(data.get('P1_Current_Temp3', 0.0)),
                float(data.get('P1_Current_PressureHose', 0.00)),
                float(data.get('P1_Current_Air_FillingTime', 0)/1000),
                float(data.get('P1_Current_Air_HoldingTime', 0)/1000),
                float(data.get('P1_Current_Air_ReleaseTime', 0)/1000),
                float(data.get('P1_Current_Oil_Start_Time', 0)/1000),
                float(data.get('P1_Current_Oil_End_Time', 0)/1000),
                int(data.get('P1_Number_Test_Times', 0)),
                float(data.get('P1_Current_PressureITV', 0.0))
            ])
            # )

            # _t("group_b", lambda: 
            self._group_b_data_filter([
                float(data.get('P2_Current_Temp1', 0.0)),
                float(data.get('P2_Current_Temp2', 0.0)),
                float(data.get('P2_Current_Temp3', 0.0)),
                float(data.get('P2_Current_PressureHose', 0.00)),
                float(data.get('P2_Current_Air_FillingTime', 0)/1000),
                float(data.get('P2_Current_Air_HoldingTime', 0)/1000),
                float(data.get('P2_Current_Air_ReleaseTime', 0)/1000),
                float(data.get('P2_Current_Oil_Start_Time', 0)/1000),
                float(data.get('P2_Current_Oil_End_Time', 0)/1000),
                int(data.get('P2_Number_Test_Times', 0)),
                float(data.get('P2_Current_PressureITV', 0.0))
            ])
            # )

            # _t("group_c", lambda: 
            self._group_c_data_filter([
                float(data.get('P3_Current_Temp1', 0.0)),
                float(data.get('P3_Current_Temp2', 0.0)),
                float(data.get('P3_Current_Temp3', 0.0)),
                float(data.get('P3_Current_PressureHose', 0.00)),
                float(data.get('P3_Current_Air_FillingTime', 0)/1000),
                float(data.get('P3_Current_Air_HoldingTime', 0)/1000),
                float(data.get('P3_Current_Air_ReleaseTime', 0)/1000),
                float(data.get('P3_Current_Oil_Start_Time', 0)/1000),
                float(data.get('P3_Current_Oil_End_Time', 0)/1000),
                int(data.get('P3_Number_Test_Times', 0)),
                float(data.get('P3_Current_PressureITV', 0.0))
            ])
            # )

            # _t("cycle_time", lambda: 
            self._set_cycle_time_unit([
                int(data.get('P1_Number_Test_Times', 0)),
                int(data.get('P2_Number_Test_Times', 0)),
                int(data.get('P3_Number_Test_Times', 0))
            ])
            # )

            # _t("at_data", lambda: 
            self._at_data_filter([
                float(data.get('P1_Current_Temp1', 0.0)),
                float(data.get('P1_Current_Temp2', 0.0)),
                float(data.get('P1_Current_Temp3', 0.0))
            ])
            # )

            # _t("bt_data", lambda: 
            self._bt_data_filter([
                float(data.get('P2_Current_Temp1', 0.0)),
                float(data.get('P2_Current_Temp2', 0.0)),
                float(data.get('P2_Current_Temp3', 0.0))
            ])
            # )

            # _t("ct_data", lambda: 
            self._ct_data_filter([
                float(data.get('P3_Current_Temp1', 0.0)),
                float(data.get('P3_Current_Temp2', 0.0)),
                float(data.get('P3_Current_Temp3', 0.0))
            ])
            # )

            # _t("alarm", lambda: 
            self._alarm_data_filter([
                bool(data.get('Bit_Alarm', False)),
                str(data.get('Alarm_Info', ""))
            ])
            # )
            if now - self._last_export_time >= 60.0:
                self._last_export_time = now
                self.auto_export_all_tables_to_excel()
        except Exception as e:
            self.logger.error("[Main]-[_data_ready]:PLC Data Processing Error: %s", e)
            
    def _set_system_data(self):
        self.init_signal = True
        
    def _init_button_obj(self, list_bool):
        self.logger.info("[Main]-[_init_button_obj]: CHECKING PLC BOOL")
        if not list_bool[1]:
            self.ui.sys_state_stacked_wid_39.setCurrentIndex(0)
            if list_bool[0]:
                self.logger.info("[Main]-[_init_button_obj]: START BOOL: 1")
                self.ui.sys_state_stacked_wid_39.setCurrentIndex(1)
                self.ui.start_stop_stacked.setCurrentIndex(1)

        if not list_bool[3]:
            if list_bool[2]:
                self.logger.info("[Main]-[_init_button_obj]: HEAT T0 BOOL: 1")
                self.ui.heat_btn_t0.click() if not self.ui.heat_btn_t0.isChecked() else None

        if list_bool[4]:
            self.logger.info("[Main]-[_init_button_obj]: HEAT A BOOL: 1")
            self.ui.heat_btn_a.click() if not self.ui.heat_btn_a.isChecked() else None
        if list_bool[5]:
            self.logger.info("[Main]-[_init_button_obj]: PRESSURE A BOOL: 1")
            self.ui.vacuum_btn_a.click() if not self.ui.vacuum_btn_a.isChecked() else None
        if list_bool[6]:
            self.logger.info("[Main]-[_init_button_obj]: OIL A BOOL: 1")
            self.ui.refuel_btn_a.click() if not self.ui.refuel_btn_a.isChecked() else None
        if list_bool[7]:
            self.logger.info("[Main]-[_init_button_obj]: CYCLE A BOOL: 1")
            self.ui.set_cycle_a_btn.click() if not self.ui.set_cycle_a_btn.isChecked() else None

        if list_bool[8]:
            self.logger.info("[Main]-[_init_button_obj]: HEAT B BOOL: 1")
            self.ui.heat_btn_b.click() if not self.ui.heat_btn_b.isChecked() else None
        if list_bool[9]:
            self.logger.info("[Main]-[_init_button_obj]: PRESSURE B BOOL: 1")
            self.ui.vacuum_btn_b.click() if not self.ui.vacuum_btn_b.isChecked() else None
        if list_bool[10]:
            self.logger.info("[Main]-[_init_button_obj]: OIL B BOOL: 1")
            self.ui.refuel_btn_b.click() if not self.ui.refuel_btn_b.isChecked() else None
        if list_bool[11]:
            self.logger.info("[Main]-[_init_button_obj]: CYCLE B BOOL: 1")
            self.ui.set_cycle_b_btn.click() if not self.ui.set_cycle_b_btn.isChecked() else None

        if list_bool[12]:
            self.logger.info("[Main]-[_init_button_obj]: HEAT C BOOL: 1")
            self.ui.heat_btn_c.click() if not self.ui.heat_btn_c.isChecked() else None
        if list_bool[13]:
            self.logger.info("[Main]-[_init_button_obj]: PRESSURE C BOOL: 1")
            self.ui.vacuum_btn_c.click() if not self.ui.vacuum_btn_c.isChecked() else None
        if list_bool[14]:
            self.logger.info("[Main]-[_init_button_obj]: OIL C BOOL: 1")
            self.ui.refuel_btn_c.click() if not self.ui.refuel_btn_c.isChecked() else None
        if list_bool[15]:
            self.logger.info("[Main]-[_init_button_obj]: CYCLE C BOOL: 1")
            self.ui.set_cycle_c_btn.click() if not self.ui.set_cycle_c_btn.isChecked() else None

    def _init_pressure_group_sv_obj(self, list_init_a, list_init_b, list_init_c, list_init_t0):
        for i in range(len(self.list_for_import_a)):
            self.list_for_import_a[i].blockSignals(True)
            # self.list_for_import_a[i].setValue(self.convert_cel_fah(list_init_a[i]))
            self.list_for_import_a[i].setValue(list_init_a[i])
            self.list_for_import_a[i].blockSignals(False)

            self.list_for_import_b[i].blockSignals(True)
            # self.list_for_import_b[i].setValue(self.convert_cel_fah(list_init_b[i]))
            self.list_for_import_b[i].setValue(list_init_b[i])
            self.list_for_import_b[i].blockSignals(False)

            self.list_for_import_c[i].blockSignals(True)
            # self.list_for_import_c[i].setValue(self.convert_cel_fah(list_init_c[i]))
            self.list_for_import_c[i].setValue(list_init_c[i])
            self.list_for_import_c[i].blockSignals(False)   

        for i in range(len(self.list_for_import_t0)):
            self.list_for_import_t0[i].blockSignals(True)
            self.list_for_import_t0[i].setValue(list_init_t0[i])
            self.list_for_import_t0[i].blockSignals(False)

    def _read_status_plc(self, connected: bool):
        if connected:
            self.ui.sys_state_stacked_wid_40.setCurrentIndex(0)
        else:
            self.ui.sys_state_stacked_wid_40.setCurrentIndex(1)
        self.plc_read_connection = connected

    def _write_status_plc(self, connected: bool):
        if connected:
            self.ui.sys_state_stacked_wid_42.setCurrentIndex(0)
        else:
            self.ui.sys_state_stacked_wid_42.setCurrentIndex(1)
        self.plc_writer_connection = connected

    def _data_group_filter(self, list_group_a_recv, list_group_b_recv, list_group_c_recv):
        if not isinstance(list_group_a_recv, list) and not isinstance(list_group_b_recv, list) and not isinstance(list_group_c_recv, list):
            return
        pv_avg = (list_group_a_recv[1] + list_group_b_recv[1] + list_group_c_recv[1]) / 3
        self.add_row_to_list_history(
            self.ui.code_display.text(),
            "Group A",
            list_group_a_recv[0],
            pv_avg,
            self.for_display_temp(list_group_a_recv[2]), 
            self.for_display_temp(list_group_a_recv[3]), 
            self.for_display_temp(list_group_a_recv[4])
        )
        self.add_row_to_list_history(
            self.ui.code_display.text(),
            "Group B",
            list_group_b_recv[0],
            pv_avg,
            self.for_display_temp(list_group_b_recv[2]), 
            self.for_display_temp(list_group_b_recv[3]), 
            self.for_display_temp(list_group_b_recv[4])
        )
        self.add_row_to_list_history(
            self.ui.code_display.text(),
            "Group C",
            list_group_c_recv[0],
            pv_avg,
            self.for_display_temp(list_group_c_recv[2]), 
            self.for_display_temp(list_group_c_recv[3]), 
            self.for_display_temp(list_group_c_recv[4])
        )

    def _t0_input_heat_filter(self, list_t0_input_recv):
        if not list_t0_input_recv[1]:
            self.ui.i_o_group_1_switch_1.setCurrentIndex(1)
            if list_t0_input_recv[0]:
                self.ui.i_o_group_1_switch_1.setCurrentIndex(0)

    def _input_data_filter(self, list_input_recv):
        for obj, value in zip(self.io_group_1_switch_obj, list_input_recv):
            obj.setCurrentIndex(value)

    def _i_o_group_3_filter(self, values):
        for i, (obj, value) in enumerate(zip(self.i_o_group_3_obj, values)):
            v = round(value, 2)
            if v != self._last_i_o_group_3[i]:
                obj.setValue(value)
                self._last_i_o_group_3[i] = v

    def _group_a_data_filter(self, list_group_a_recv):
        avg_a = round(self.for_display_temp(
            (list_group_a_recv[0]+list_group_a_recv[1]+list_group_a_recv[2])/3), 2)
        if avg_a != self._last_group_a_avg:
            self.temp_pv_obj[1].setValue(avg_a)
            self._last_group_a_avg = avg_a

        for i, val_a in enumerate(list_group_a_recv):
            v = round(val_a if i >= 3 else self.for_display_temp(val_a), 2)
            if v != self._last_group_a[i]:
                self.pressure_a_pv_obj[i].setValue(v)
                self._last_group_a[i] = v

    def _group_b_data_filter(self, list_group_b_recv):
        avg_b = round(self.for_display_temp(
            (list_group_b_recv[0]+list_group_b_recv[1]+list_group_b_recv[2])/3), 2)
        if avg_b != self._last_group_b_avg:
            self.temp_pv_obj[2].setValue(avg_b)
            self._last_group_b_avg = avg_b

        for i, val_b in enumerate(list_group_b_recv):
            v = round(val_b if i >= 3 else self.for_display_temp(val_b), 2)
            if v != self._last_group_b[i]:
                self.pressure_b_pv_obj[i].setValue(v)
                self._last_group_b[i] = v

    def _group_c_data_filter(self, list_group_c_recv):
        avg_c = round(self.for_display_temp(
            (list_group_c_recv[0]+list_group_c_recv[1]+list_group_c_recv[2])/3), 2)
        if avg_c != self._last_group_c_avg:
            self.temp_pv_obj[3].setValue(avg_c)
            self._last_group_c_avg = avg_c

        for i, val_c in enumerate(list_group_c_recv):
            v = round(val_c if i >= 3 else self.for_display_temp(val_c), 2)
            if v != self._last_group_c[i]:
                self.pressure_c_pv_obj[i].setValue(v)
                self._last_group_c[i] = v

    def _t0_data_filter(self, list_group_t0_recv):
        v = round(self.for_display_temp(list_group_t0_recv[1]), 2)
        if v != self._last_t0_pv:
            self.temp_pv_obj[0].setValue(v)
            self._last_t0_pv = v

    def _set_cycle_time_unit(self, list_cycle_recv):
        for i, (displ_obj, val) in enumerate(zip(
            [self.ui.cycle_a_displ_3,  self.ui.cycle_b_displ_3,  self.ui.cycle_c_displ_3],
            list_cycle_recv
        )):
            if val != self._last_cycle[i]:
                displ_obj.setValue(val)
                self._last_cycle[i] = val

    def _at_data_filter(self, list_group_at_recv):
        v = round(self.for_display_temp(
            (list_group_at_recv[0]+list_group_at_recv[1]+list_group_at_recv[2])/3), 2)
        if v != self._last_at:
            self.ui.at_pv.setValue(v)
            self._last_at = v

    def _bt_data_filter(self, list_group_bt_recv):
        v = round(self.for_display_temp(
            (list_group_bt_recv[0]+list_group_bt_recv[1]+list_group_bt_recv[2])/3), 2)
        if v != self._last_bt:
            self.ui.bt_pv.setValue(v)
            self._last_bt = v

    def _ct_data_filter(self, list_group_ct_recv):
        v = round(self.for_display_temp(
            (list_group_ct_recv[0]+list_group_ct_recv[1]+list_group_ct_recv[2])/3), 2)
        if v != self._last_ct:
            self.ui.ct_pv.setValue(v)
            self._last_ct = v

    def _alarm_data_filter(self, alarm_recv):
        if alarm_recv[0]:
            self.ui.error_display.setText(alarm_recv[1])
            # print(alarm_recv[1])

    def on_heat_btn_clicked(self, channel: str, checked: bool):
        simulator = self.thread_dict.get("data_simulator")
        if simulator is None:
            return

        if checked:
            if   channel == "A":
                pressure_sv = self.pressure_a_sv_obj[4].value() 
                temp_sv     = self.pressure_a_sv_obj[0].value() if self.pressure_a_sv_obj[0].value() > self.default_temp_room else self.default_temp_room
            elif channel == "B":
                pressure_sv = self.pressure_b_sv_obj[4].value()
                temp_sv     = self.pressure_b_sv_obj[0].value() if self.pressure_b_sv_obj[0].value() > self.default_temp_room else self.default_temp_room
            elif channel == "C":
                pressure_sv = self.pressure_c_sv_obj[4].value()
                temp_sv     = self.pressure_c_sv_obj[0].value() if self.pressure_c_sv_obj[0].value() > self.default_temp_room else self.default_temp_room

            simulator.set_channel_active(channel, pressure_sv, temp_sv)
        else:
            simulator.set_channel_active(channel, 0.0, self.default_temp_room)  # Reset về mặc định

    def heating_btn(self, channel: str, checked: bool, btn=None):
        if not self.plc_writer_connection and not self.init_signal:
            ltmessage.error(self, "Error", "PLC Writer not connected!")
            if btn is not None:
                btn.blockSignals(True)   # Chặn signal để tránh gọi đệ quy
                btn.setChecked(False)
                btn.blockSignals(False)
            return
        else:
            if channel == "A":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P1_Start_Heat", True)
                    self.disable_heat_group(channel, False)
                    self.logger.info("[Main]-[heating_btn]: Group A Heating On!")
                    # ltmessage.information(self, "Heating", "Group A Heating On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P1_Start_Heat", False)
                    self.disable_heat_group(channel, True)
                    self.logger.info("[Main]-[heating_btn]: Group A Heating Off!")
                return
            if channel == "B":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P2_Start_Heat", True)
                    self.disable_heat_group(channel, False)
                    self.logger.info("[Main]-[heating_btn]: Group B Heating On!")
                    # ltmessage.information(self, "Heating", "Group B Heating On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P2_Start_Heat", False)
                    self.disable_heat_group(channel, True)
                    self.logger.info("[Main]-[heating_btn]: Group B Heating Off!")
                return
            if channel == "C":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P3_Start_Heat", True)
                    self.disable_heat_group(channel, False)
                    self.logger.info("[Main]-[heating_btn]: Group C Heating On!")
                    # ltmessage.information(self, "Heating", "Group C Heating On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P3_Start_Heat", False)
                    self.disable_heat_group(channel, True)
                    self.logger.info("[Main]-[heating_btn]: Group C Heating Off!")
                return
            if channel == "T0":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("T0_Start_Heat", True)
                    self.disable_heat_group(channel, False)
                    self.logger.info("[Main]-[heating_btn]: T0 Heating On!")
                    # ltmessage.information(self, "Heating", "T0 Heating On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("T0_Stop_Heat", True)
                        QTimer.singleShot(100,  lambda: self.plc_writer_worker.write_bool.emit("T0_Start_Heat", False))
                        QTimer.singleShot(200,  lambda: self.plc_writer_worker.write_bool.emit("T0_Stop_Heat", False))
                    self.disable_heat_group(channel, True)
                    self.logger.info("[Main]-[heating_btn]: T0 Heating Off!")
                return

    def pumping_btn(self, channel: str, checked: bool, btn=None):
        if not self.plc_writer_connection and not self.init_signal:
            ltmessage.error(self, "Error", "PLC Writer not connected!")
            if btn is not None:
                btn.blockSignals(True)   # Chặn signal để tránh gọi đệ quy
                btn.setChecked(False)
                btn.blockSignals(False)
            return
        else:
            if channel == "A":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P1_Start_Pressure", True)
                    self.disable_pressure_group(channel, False)
                    self.logger.info("[Main]-[pumping_btn]: Group A Pressure On!")
                    # ltmessage.information(self, "Pumping", "Group A Pressure On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P1_Start_Pressure", False)
                    self.disable_pressure_group(channel, True)
                    self.logger.info("[Main]-[pumping_btn]: Group A Pressure Off!")

            elif channel == "B":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P2_Start_Pressure", True)
                    self.disable_pressure_group(channel, False)
                    self.logger.info("[Main]-[pumping_btn]: Group B Pressure On!")
                    # ltmessage.information(self, "Pumping", "Group B Pressure On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P2_Start_Pressure", False)
                    self.disable_pressure_group(channel, True)
                    self.logger.info("[Main]-[pumping_btn]: Group B Pressure Off!")

            elif channel == "C":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P3_Start_Pressure", True)
                    self.disable_pressure_group(channel, False)
                    self.logger.info("[Main]-[pumping_btn]: Group C Pressure On!")
                    # ltmessage.information(self, "Pumping", "Group C Pressure On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P3_Start_Pressure", False)
                    self.disable_pressure_group(channel, True)
                    self.logger.info("[Main]-[pumping_btn]: Group C Pressure Off!")
            return

    def fill_oil_btn(self, channel: str, checked: bool, btn=None):
        if not self.plc_writer_connection and not self.init_signal:
            ltmessage.error(self, "Error", "PLC Writer not connected!")
            if btn is not None:
                btn.blockSignals(True)   # Chặn signal để tránh gọi đệ quy
                btn.setChecked(False)
                btn.blockSignals(False)
            return
        else:
            if channel == "A":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P1_Start_Oil", True)
                    self.disable_oil_group(channel, False)
                    self.logger.info("[Main]-[fill_oil_btn]: Group A Oil Filling On!")
                    # ltmessage.information(self, "Oil Fill", "Group A Oil Filling On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P1_Start_Oil", False)
                    self.disable_oil_group(channel, True)
                    self.logger.info("[Main]-[fill_oil_btn]: Group A Oil Filling Off!")

            if channel == "B":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P2_Start_Oil", True)
                    self.disable_oil_group(channel, False)
                    self.logger.info("[Main]-[fill_oil_btn]: Group B Oil Filling On!")
                    # ltmessage.information(self, "Oil Fill", "Group B Oil Filling On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P2_Start_Oil", False)
                    self.disable_oil_group(channel, True)
                    self.logger.info("[Main]-[fill_oil_btn]: Group B Oil Filling Off!")

            if channel == "C":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P3_Start_Oil", True)
                    self.disable_oil_group(channel, False)
                    self.logger.info("[Main]-[fill_oil_btn]: Group C Oil Filling On!")
                    # ltmessage.information(self, "Oil Fill", "Group C Oil Filling On!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P3_Start_Oil", False)
                    self.disable_oil_group(channel, True)
                    self.logger.info("[Main]-[fill_oil_btn]: Group C Oil Filling Off!")

    def cycle_loop_btn(self, channel: str, checked: bool, btn=None):
        if not self.plc_writer_connection and not self.init_signal:
            ltmessage.error(self, "Error", "PLC Writer not connected!")
            if btn is not None:
                btn.blockSignals(True)   # Chặn signal để tránh gọi đệ quy
                btn.setChecked(False)
                btn.blockSignals(False)
            return
        else:
            if channel == "A":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P1_BitCountTimes", True)
                    self.ui.pressure_sv_a_11.setEnabled(False)
                    self.logger.info("[Main]-[cycle_loop_btn]: Group A Auto Repeat Off!")
                    # ltmessage.information(self, "Set Cycle A", "Group A Auto Repeat!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P1_BitCountTimes", False)
                    self.ui.pressure_sv_a_11.setEnabled(True)
                    self.logger.info("[Main]-[cycle_loop_btn]: Group A Auto Repeat On!")
                return
            if channel == "B":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P2_BitCountTimes", True)
                    self.ui.pressure_sv_b_11.setEnabled(False)
                    self.logger.info("[Main]-[cycle_loop_btn]: Group B Auto Repeat Off!")
                    # ltmessage.information(self, "Set Cycle B", "Group B Auto Repeat!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P2_BitCountTimes", False)
                    self.ui.pressure_sv_b_11.setEnabled(True)
                    self.logger.info("[Main]-[cycle_loop_btn]: Group B Auto Repeat On!")
                return
            if channel == "C":
                if checked:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P3_BitCountTimes", True)
                    self.ui.pressure_sv_c_11.setEnabled(False)
                    self.logger.info("[Main]-[cycle_loop_btn]: Group C Auto Repeat Off!")
                    # ltmessage.information(self, "Set Cycle C", "Group C Auto Repeat!")
                else:
                    if not self.init_signal:
                        self.plc_writer_worker.write_bool.emit("P3_BitCountTimes", False)
                    self.ui.pressure_sv_c_11.setEnabled(True)
                    self.logger.info("[Main]-[cycle_loop_btn]: Group C Auto Repeat On!")
                return

    def start_stop_btn(self, btn=None):
        if not self.plc_writer_connection and not self.init_signal:
            ltmessage.error(self, "Error", "PLC Writer not connected!")
            if btn is not None:
                btn.blockSignals(True)   # Chặn signal để tránh gọi đệ quy
                btn.setChecked(False)
                btn.blockSignals(False)
            return
        if self.ui.start_stop_stacked.currentIndex() == 0:
            self.ui.sys_state_stacked_wid_39.setCurrentIndex(1)
            self.ui.start_stop_stacked.setCurrentIndex(1)
            self.plc_writer_worker.write_bool.emit("START", True)
            self.logger.info("[Main]-[start_stop_btn]: System On")
            # QTimer.singleShot(250, lambda: self.plc_writer_worker.write_bool.emit("START", False))
            # ltmessage.information(self, "Strike Machine", "System On!")
        elif self.ui.start_stop_stacked.currentIndex() == 1:
            self.ui.sys_state_stacked_wid_39.setCurrentIndex(0)
            self.ui.start_stop_stacked.setCurrentIndex(0)
            self.plc_writer_worker.write_bool.emit("STOP", True)
            QTimer.singleShot(100, lambda: self.plc_writer_worker.write_bool.emit("START", False))
            QTimer.singleShot(200, lambda: self.plc_writer_worker.write_bool.emit("STOP", False))
            self.logger.info("[Main]-[start_stop_btn]: System Off")

    def clear_data_btn(self):
        reply = ltmessage.question(
            self, "Clear Data", "Set all SV to 0?"
        )
        if reply == ltmessage.Yes:
            try:
                self.ui.code_display.setText("")
                for i in range(len(self.pressure_a_sv_obj)):
                    self.pressure_a_sv_obj[i].blockSignals(False)
                    self.pressure_b_sv_obj[i].blockSignals(False)
                    self.pressure_c_sv_obj[i].blockSignals(False)                    
                    self.pressure_a_sv_obj[i].setValue(0)
                    self.pressure_b_sv_obj[i].setValue(0)
                    self.pressure_c_sv_obj[i].setValue(0)
                    self.pressure_a_sv_obj[i].blockSignals(True)
                    self.pressure_b_sv_obj[i].blockSignals(True)
                    self.pressure_c_sv_obj[i].blockSignals(True)       
                for i in range(len(self.temp_sv_obj)):
                    self.temp_sv_obj[i].blockSignals(False)
                    self.temp_sv_obj[i].setValue(0)
                    self.temp_sv_obj[i].blockSignals(True)
                for i in range(len(self.temp_h_alm_obj)):
                    self.temp_h_alm_obj[i].blockSignals(False)
                    self.temp_h_alm_obj[i].setValue(0)
                    self.temp_h_alm_obj[i].blockSignals(True)

                for i in range(len(self.temp_l_alm_obj)):
                    self.temp_l_alm_obj[i].blockSignals(False)
                    self.temp_l_alm_obj[i].setValue(0)
                    self.temp_l_alm_obj[i].blockSignals(True)
                for i in range(len(self.temp_offset_obj)):
                    self.temp_offset_obj[i].blockSignals(False)
                    self.temp_offset_obj[i].setValue(0)
                    self.temp_offset_obj[i].blockSignals(True)
                items_a = [
                    self.plc_writer_worker.get_item("P1_CountTimes", self.list_for_import_a[0].value()),
                    self.plc_writer_worker.get_item("P1_Oil_Start_Time", self.list_for_import_a[1].value()),
                    self.plc_writer_worker.get_item("P1_Oil_End_Time", self.list_for_import_a[2].value()),
                    self.plc_writer_worker.get_item("P1_Air_FillingTime", self.list_for_import_a[3].value()),
                    self.plc_writer_worker.get_item("P1_Air_HoldingTime", self.list_for_import_a[4].value()),
                    self.plc_writer_worker.get_item("P1_Air_ReleaseTime", self.list_for_import_a[5].value()),
                    self.plc_writer_worker.get_item("P1_PressureSetting", self.list_for_import_a[6].value()),
                    self.plc_writer_worker.get_item("P1_TemperatureSetting", self.list_for_import_a[7].value()),
                    self.plc_writer_worker.get_item("P1_TempLimitHIGH", self.list_for_import_a[8].value()),
                    self.plc_writer_worker.get_item("P1_TempLimitLOW", self.list_for_import_a[9].value()),
                    self.plc_writer_worker.get_item("P1_Temp1Offset", self.list_for_import_a[10].value()),
                    self.plc_writer_worker.get_item("P1_Temp2Offset", self.list_for_import_a[11].value()),
                    self.plc_writer_worker.get_item("P1_Temp3Offset", self.list_for_import_a[12].value())
                ]
                items_b = [
                    self.plc_writer_worker.get_item("P2_CountTimes", self.list_for_import_b[0].value()),
                    self.plc_writer_worker.get_item("P2_Oil_Start_Time", self.list_for_import_b[1].value()),
                    self.plc_writer_worker.get_item("P2_Oil_End_Time", self.list_for_import_b[2].value()),
                    self.plc_writer_worker.get_item("P2_Air_FillingTime", self.list_for_import_b[3].value()),
                    self.plc_writer_worker.get_item("P2_Air_HoldingTime", self.list_for_import_b[4].value()),
                    self.plc_writer_worker.get_item("P2_Air_ReleaseTime", self.list_for_import_b[5].value()),
                    self.plc_writer_worker.get_item("P2_PressureSetting", self.list_for_import_b[6].value()),
                    self.plc_writer_worker.get_item("P2_TemperatureSetting", self.list_for_import_b[7].value()),
                    self.plc_writer_worker.get_item("P2_TempLimitHIGH", self.list_for_import_b[8].value()),
                    self.plc_writer_worker.get_item("P2_TempLimitLOW", self.list_for_import_b[9].value()),
                    self.plc_writer_worker.get_item("P2_Temp1Offset", self.list_for_import_b[10].value()),
                    self.plc_writer_worker.get_item("P2_Temp2Offset", self.list_for_import_b[11].value()),
                    self.plc_writer_worker.get_item("P2_Temp3Offset", self.list_for_import_b[12].value())
                ]
                items_c = [
                    self.plc_writer_worker.get_item("P3_CountTimes", self.list_for_import_c[0].value()),
                    self.plc_writer_worker.get_item("P3_Oil_Start_Time", self.list_for_import_c[1].value()),
                    self.plc_writer_worker.get_item("P3_Oil_End_Time", self.list_for_import_c[2].value()),
                    self.plc_writer_worker.get_item("P3_Air_FillingTime", self.list_for_import_c[3].value()),
                    self.plc_writer_worker.get_item("P3_Air_HoldingTime", self.list_for_import_c[4].value()),
                    self.plc_writer_worker.get_item("P3_Air_ReleaseTime", self.list_for_import_c[5].value()),
                    self.plc_writer_worker.get_item("P3_PressureSetting", self.list_for_import_c[6].value()),
                    self.plc_writer_worker.get_item("P3_TemperatureSetting", self.list_for_import_c[7].value()),
                    self.plc_writer_worker.get_item("P3_TempLimitHIGH", self.list_for_import_c[8].value()),
                    self.plc_writer_worker.get_item("P3_TempLimitLOW", self.list_for_import_c[9].value()),
                    self.plc_writer_worker.get_item("P3_Temp1Offset", self.list_for_import_c[10].value()),
                    self.plc_writer_worker.get_item("P3_Temp2Offset", self.list_for_import_c[11].value()),
                    self.plc_writer_worker.get_item("P3_Temp3Offset", self.list_for_import_c[12].value())
                ]
                items_t0 = [
                    self.plc_writer_worker.get_item("T0_TemperatureSetting", self.list_for_import_t0[0].value()),
                    self.plc_writer_worker.get_item("T0_TempLimitHIGH", self.list_for_import_t0[1].value()),
                    self.plc_writer_worker.get_item("T0_TempLimitLOW", self.list_for_import_t0[2].value()),
                    self.plc_writer_worker.get_item("T0_TempOffset", self.list_for_import_t0[3].value())
                ]
                self.plc_writer_worker.write_multi.emit(items_a)
                self.plc_writer_worker.write_multi.emit(items_b)
                self.plc_writer_worker.write_multi.emit(items_c)
                self.plc_writer_worker.write_multi.emit(items_t0)
            except Exception as e:
                ltmessage.error(self, "Error", f"Failed to clear data: {e}")

    def new_data_btn(self):
        stk_mch_file = Path(self.stk_mch_folder)/ "Setting File" 
        # print("Default path for data file:", stk_mch_file)s
        file_str, _ = QFileDialog.getOpenFileName(
            self,
            "Select Group Data File",
            str(stk_mch_file),
            "Excel Files (*.xlsx *.xls)"
        )
        if not file_str:  # User cancel
            return
        path = Path(file_str)

        df = pd.read_excel(path, sheet_name='Sheet1', header=None)
        self.ui.code_display.setText(str(df.iloc[0][0]).strip() if pd.notna(df.iloc[0][0]) else "")
        for i in range(2, len(df)): # Bắt đầu từ hàng 3
            column = df.iloc[i]
            name_raw = str(column[0]).strip() if pd.notna(column[0]) else ""
            if name_raw == "" or name_raw.lower() == "nan":
                break
            # Cột B: Group A
            try:
                value_a_raw = str(column[1]).strip()
                value_a = float(value_a_raw)
            except:
                value_a = 0
            self.list_for_import_a[i-2].blockSignals(True)
            self.list_for_import_a[i-2].setValue(value_a)
            
            # Cột C: Group B
            try:
                value_b_raw = str(column[2]).strip()
                value_b = float(value_b_raw)
            except:
                value_b = 0
            self.list_for_import_c[i-2].blockSignals(True)
            self.list_for_import_b[i-2].setValue(value_b)
            
            # Cột D: Group C
            try:
                value_c_raw = str(column[3]).strip()
                value_c = float(value_c_raw)
            except:
                value_c = 0
            self.list_for_import_c[i-2].blockSignals(True)
            self.list_for_import_c[i-2].setValue(value_c)

            if i >= 9  and i <=12:
                # Cột E: Group T0
                try:
                    value_t0_raw = str(column[4]).strip()
                    value_t0 = float(value_t0_raw)
                except:
                    value_t0 = 0
                self.list_for_import_t0[i-9].blockSignals(True)
                self.list_for_import_t0[i-9].setValue(value_t0)

        self.ui.at_sv.blockSignals(True)
        self.ui.at_sv.setValue(self.ui.pressure_sv_a_1.value())
        self.ui.bt_sv.blockSignals(True)
        self.ui.bt_sv.setValue(self.ui.pressure_sv_b_1.value())
        self.ui.ct_sv.blockSignals(True)
        self.ui.ct_sv.setValue(self.ui.pressure_sv_c_1.value())
        items_a = [
            self.plc_writer_worker.get_item("P1_CountTimes", self.list_for_import_a[0].value()),
            self.plc_writer_worker.get_item("P1_Oil_Start_Time", self.list_for_import_a[1].value()),
            self.plc_writer_worker.get_item("P1_Oil_End_Time", self.list_for_import_a[2].value()),
            self.plc_writer_worker.get_item("P1_Air_FillingTime", self.list_for_import_a[3].value()),
            self.plc_writer_worker.get_item("P1_Air_HoldingTime", self.list_for_import_a[4].value()),
            self.plc_writer_worker.get_item("P1_Air_ReleaseTime", self.list_for_import_a[5].value()),
            self.plc_writer_worker.get_item("P1_PressureSetting", self.list_for_import_a[6].value()),
            self.plc_writer_worker.get_item("P1_TemperatureSetting", self.list_for_import_a[7].value()),
            self.plc_writer_worker.get_item("P1_TempLimitHIGH", self.list_for_import_a[8].value()),
            self.plc_writer_worker.get_item("P1_TempLimitLOW", self.list_for_import_a[9].value()),
            self.plc_writer_worker.get_item("P1_Temp1Offset", self.list_for_import_a[10].value()),
            self.plc_writer_worker.get_item("P1_Temp2Offset", self.list_for_import_a[11].value()),
            self.plc_writer_worker.get_item("P1_Temp3Offset", self.list_for_import_a[12].value())
        ]
        items_b = [
            self.plc_writer_worker.get_item("P2_CountTimes", self.list_for_import_b[0].value()),
            self.plc_writer_worker.get_item("P2_Oil_Start_Time", self.list_for_import_b[1].value()),
            self.plc_writer_worker.get_item("P2_Oil_End_Time", self.list_for_import_b[2].value()),
            self.plc_writer_worker.get_item("P2_Air_FillingTime", self.list_for_import_b[3].value()),
            self.plc_writer_worker.get_item("P2_Air_HoldingTime", self.list_for_import_b[4].value()),
            self.plc_writer_worker.get_item("P2_Air_ReleaseTime", self.list_for_import_b[5].value()),
            self.plc_writer_worker.get_item("P2_PressureSetting", self.list_for_import_b[6].value()),
            self.plc_writer_worker.get_item("P2_TemperatureSetting", self.list_for_import_b[7].value()),
            self.plc_writer_worker.get_item("P2_TempLimitHIGH", self.list_for_import_b[8].value()),
            self.plc_writer_worker.get_item("P2_TempLimitLOW", self.list_for_import_b[9].value()),
            self.plc_writer_worker.get_item("P2_Temp1Offset", self.list_for_import_b[10].value()),
            self.plc_writer_worker.get_item("P2_Temp2Offset", self.list_for_import_b[11].value()),
            self.plc_writer_worker.get_item("P2_Temp3Offset", self.list_for_import_b[12].value())
        ]
        items_c = [
            self.plc_writer_worker.get_item("P3_CountTimes", self.list_for_import_c[0].value()),
            self.plc_writer_worker.get_item("P3_Oil_Start_Time", self.list_for_import_c[1].value()),
            self.plc_writer_worker.get_item("P3_Oil_End_Time", self.list_for_import_c[2].value()),
            self.plc_writer_worker.get_item("P3_Air_FillingTime", self.list_for_import_c[3].value()),
            self.plc_writer_worker.get_item("P3_Air_HoldingTime", self.list_for_import_c[4].value()),
            self.plc_writer_worker.get_item("P3_Air_ReleaseTime", self.list_for_import_c[5].value()),
            self.plc_writer_worker.get_item("P3_PressureSetting", self.list_for_import_c[6].value()),
            self.plc_writer_worker.get_item("P3_TemperatureSetting", self.list_for_import_c[7].value()),
            self.plc_writer_worker.get_item("P3_TempLimitHIGH", self.list_for_import_c[8].value()),
            self.plc_writer_worker.get_item("P3_TempLimitLOW", self.list_for_import_c[9].value()),
            self.plc_writer_worker.get_item("P3_Temp1Offset", self.list_for_import_c[10].value()),
            self.plc_writer_worker.get_item("P3_Temp2Offset", self.list_for_import_c[11].value()),
            self.plc_writer_worker.get_item("P3_Temp3Offset", self.list_for_import_c[12].value())
        ]
        items_t0 = [
            self.plc_writer_worker.get_item("T0_TemperatureSetting", self.list_for_import_t0[0].value()),
            self.plc_writer_worker.get_item("T0_TempLimitHIGH", self.list_for_import_t0[1].value()),
            self.plc_writer_worker.get_item("T0_TempLimitLOW", self.list_for_import_t0[2].value()),
            self.plc_writer_worker.get_item("T0_TempOffset", self.list_for_import_t0[3].value())
        ]
        self.plc_writer_worker.write_multi.emit(items_a)
        self.plc_writer_worker.write_multi.emit(items_b)
        self.plc_writer_worker.write_multi.emit(items_c)
        self.plc_writer_worker.write_multi.emit(items_t0)
        for i in range(len(self.list_for_import_a)):
            self.list_for_import_a[i].blockSignals(False)
            self.list_for_import_b[i].blockSignals(False)
            self.list_for_import_c[i].blockSignals(False)
        for i in range(len(self.list_for_import_t0)):
            self.list_for_import_t0[i].blockSignals(False)
        
        self.ui.at_sv.blockSignals(False)
        self.ui.bt_sv.blockSignals(False)
        self.ui.ct_sv.blockSignals(False)
        self.ui.set_cycle_a_btn.click()
        self.ui.set_cycle_b_btn.click()
        self.ui.set_cycle_c_btn.click()

    def set_language_en(self):
        self._app.removeTranslator(self._translator)
        self._current_lang = "en"

    def set_language_cn(self):
        self._app.removeTranslator(self._translator)
        self._translator.load(resource_path("tech_link_theme_cn.qm"))
        self._app.installTranslator(self._translator)
        self._current_lang = "cn"
        
    def changeEvent(self, event):
        super().changeEvent(event)
        if event.type() == QEvent.Type.LanguageChange:
            ip_text = self.ui.plc_ip_address_edit.text()
            db_text = self.ui.db_file_path_edit.text()
            self.ui.retranslateUi(self)
            charts = [self.chart_temp, self.chart_pressure_a,
                self.chart_pressure_b, self.chart_pressure_c]
            if self._current_lang == "cn":
                self.ui.plc_ip_address_edit.setPlaceholderText("请输入IP地址: 172.16.100.***")
                self.ui.db_file_path_edit.setPlaceholderText("输入路径文件夹")
                
                self.chart_temp.btn_setting.setText("烤箱")
                self.chart_temp.plot.setLabel("left", "温度 (°C)") if self._current_unit == 0 else self.chart_temp.plot.setLabel("left", "温度 (°F)")
                
                self.chart_pressure_a.btn_setting.setText("A组")
                self.chart_pressure_a.plot.setLabel("left", "温度 (°C)") if self._current_unit == 0 else self.chart_pressure_a.plot.setLabel("left", "温度 (°F)")
                self.chart_pressure_a.plot.plotItem.axes["right"]["item"].setLabel("压力 (bar)")
                
                self.chart_pressure_b.btn_setting.setText("B组")
                self.chart_pressure_b.plot.setLabel("left", "温度 (°C)") if self._current_unit == 0 else self.chart_pressure_b.plot.setLabel("left", "温度 (°F)")
                self.chart_pressure_b.plot.plotItem.axes["right"]["item"].setLabel("压力 (bar)")
                
                self.chart_pressure_c.btn_setting.setText("C组")
                self.chart_pressure_c.plot.setLabel("left", "温度 (°C)") if self._current_unit == 0 else self.chart_pressure_c.plot.setLabel("left", "温度 (°F)")
                self.chart_pressure_c.plot.plotItem.axes["right"]["item"].setLabel("压力 (bar)")

            elif self._current_lang == "en":
                self.ui.plc_ip_address_edit.setPlaceholderText("Enter IP Address: 172.16.100.***")
                self.ui.db_file_path_edit.setPlaceholderText("Enter Path Folder")
                
                self.chart_temp.btn_setting.setText("Oven")
                self.chart_temp.plot.setLabel("left", "Temperature (°C)") if self._current_unit == 0 else self.chart_temp.plot.setLabel("left", "Temperature (°F)")
                
                self.chart_pressure_a.btn_setting.setText("Group A")
                self.chart_pressure_a.plot.setLabel("left", "Temperature (°C)") if self._current_unit == 0 else self.chart_pressure_a.plot.setLabel("left", "Temperature (°F)")
                self.chart_pressure_a.plot.plotItem.axes["right"]["item"].setLabel("Pressure (bar)")
                
                self.chart_pressure_b.btn_setting.setText("Group B")
                self.chart_pressure_b.plot.setLabel("left", "Temperature (°C)") if self._current_unit == 0 else self.chart_pressure_b.plot.setLabel("left", "Temperature (°F)")
                self.chart_pressure_b.plot.plotItem.axes["right"]["item"].setLabel("Pressure (bar)")
                
                self.chart_pressure_c.btn_setting.setText("Group C")
                self.chart_pressure_c.plot.setLabel("left", "Temperature (°C)") if self._current_unit == 0 else self.chart_pressure_c.plot.setLabel("left", "Temperature (°F)")
                self.chart_pressure_c.plot.plotItem.axes["right"]["item"].setLabel("Pressure (bar)")
                
            self.ui.plc_ip_address_edit.setText(ip_text)
            self.ui.db_file_path_edit.setText(db_text)

    def _set_cur_unit(self):
        index = self.ui.temp_unit_selection_combox.currentIndex()
        for i in range(len(self.cel_fah_change)):
            self.cel_fah_change[i].setCurrentIndex(index)
        self._current_unit = index
        self._set_sv_widget_pressure_obj()
        self._set_sv_widget_temperature_obj()
        self._set_chart_unit()

    def _set_sv_widget_pressure_obj(self):
        self.pressure_a_sv_obj[0].blockSignals(True)
        self.pressure_a_sv_obj[0].setValue(self.convert_cel_fah(self.pressure_a_sv_obj[0].value()))
        self.pressure_a_sv_obj[0].blockSignals(False)
        self.pressure_b_sv_obj[0].blockSignals(True)
        self.pressure_b_sv_obj[0].setValue(self.convert_cel_fah(self.pressure_b_sv_obj[0].value()))
        self.pressure_b_sv_obj[0].blockSignals(False)
        self.pressure_c_sv_obj[0].blockSignals(True)
        self.pressure_c_sv_obj[0].setValue(self.convert_cel_fah(self.pressure_c_sv_obj[0].value()))
        self.pressure_c_sv_obj[0].blockSignals(False)

    def _set_sv_widget_temperature_obj(self):
        for i in range(4):
            self.temp_sv_obj[i].blockSignals(True)
            self.temp_sv_obj[i].setValue(self.convert_cel_fah(self.temp_sv_obj[i].value()))
            self.temp_sv_obj[i].blockSignals(False)

            self.temp_h_alm_obj[i].blockSignals(True)
            self.temp_h_alm_obj[i].setValue(self.convert_cel_fah(self.temp_h_alm_obj[i].value()))
            self.temp_h_alm_obj[i].blockSignals(False)

            self.temp_l_alm_obj[i].blockSignals(True)
            self.temp_l_alm_obj[i].setValue(self.convert_cel_fah(self.temp_l_alm_obj[i].value()))
            self.temp_l_alm_obj[i].blockSignals(False)

        for i in range(len(self.temp_offset_obj)):
            self.temp_offset_obj[i].blockSignals(True)
            self.temp_offset_obj[i].setValue(self.convert_cel_fah(self.temp_offset_obj[i].value()))
            self.temp_offset_obj[i].blockSignals(False)

    def _set_chart_unit(self):
        if self._current_unit == 0:
            self.chart_temp.set_temp_label("Temperature (°C)", unit="°C") if self._current_lang == "en" else self.chart_temp.set_temp_label("温度 (°C)", unit="°C")
            self.chart_pressure_a.set_temp_label("Temperature (°C)", unit="°C") if self._current_lang == "en" else self.chart_pressure_a.set_temp_label("温度 (°C)", unit="°C")
            self.chart_pressure_b.set_temp_label("Temperature (°C)", unit="°C") if self._current_lang == "en" else self.chart_pressure_b.set_temp_label("温度 (°C)", unit="°C")
            self.chart_pressure_c.set_temp_label("Temperature (°C)", unit="°C") if self._current_lang == "en" else self.chart_pressure_c.set_temp_label("温度 (°C)", unit="°C")

        elif self._current_unit == 1:
            self.chart_temp.set_temp_label("Temperature (°F)", unit="°F") if self._current_lang == "en" else self.chart_temp.set_temp_label("温度 (°F)", unit="°F")
            self.chart_pressure_a.set_temp_label("Temperature (°F)", unit="°F") if self._current_lang == "en" else self.chart_pressure_a.set_temp_label("温度 (°F)", unit="°F")
            self.chart_pressure_b.set_temp_label("Temperature (°F)", unit="°F") if self._current_lang == "en" else self.chart_pressure_b.set_temp_label("温度 (°F)", unit="°F")
            self.chart_pressure_c.set_temp_label("Temperature (°F)", unit="°F") if self._current_lang == "en" else self.chart_pressure_c.set_temp_label("温度 (°F)", unit="°F")
    
    def for_display_temp(self, celsius):
        if self._current_unit == 1:
            return celsius * 9/5 + 32
        return celsius

    # def _update_alarm_state(self, )

    def convert_cel_fah(self, temp):
        if self._current_unit == 1:
            return temp * 9/5 + 32
        elif self._current_unit == 0:
            return (temp - 32) * 5/9

    def cal_fah_to_cel(self, temp):
        if self._current_unit == 1:
            return (temp - 32) * 5/9
        return temp
    
    def cal_sec_to_msec(self, time_sec):
        return time_sec * 1000

    def apply_group_to_ui(self, group_dict: dict, widgets: list):
        for widget, value in zip(widgets, group_dict.values()):
            widget.blockSignals(True)
            try:
                if hasattr(widget, 'setValue'):
                    widget.setValue(float(value))
            finally:
                widget.blockSignals(False)
                
    def update_chart_temp(self):
        self.chart_temp.append_data([
            self.ui.t0_sv.value(),
            self.ui.t0_pv.value()
        ])

    def update_chart_pressure_a(self):
        group_a_values = [self.ui.pressure_sv_a_1.value(),
                        self.ui.pressure_pv_a_2.value(),
                        self.ui.pressure_pv_a_3.value(),
                        self.ui.pressure_pv_a_4.value()]
        group_a_press_values = [self.ui.pressure_sv_a_5.value(),
                                self.ui.pressure_pv_a_5.value()]
        self.chart_pressure_a.append_data(group_a_values, group_a_press_values)

    def update_chart_pressure_b(self):
        group_b_values = [self.ui.pressure_sv_b_1.value(),
                        self.ui.pressure_pv_b_2.value(),
                        self.ui.pressure_pv_b_3.value(),
                        self.ui.pressure_pv_b_4.value()]
        group_b_press_values = [self.ui.pressure_sv_b_5.value(),
                                self.ui.pressure_pv_b_5.value()]
        self.chart_pressure_b.append_data(group_b_values, group_b_press_values)

    def update_chart_pressure_c(self):
        group_c_values = [self.ui.pressure_sv_c_1.value(),
                        self.ui.pressure_pv_c_2.value(),
                        self.ui.pressure_pv_c_3.value(),
                        self.ui.pressure_pv_c_4.value()]
        group_c_press_values = [self.ui.pressure_sv_c_5.value(),
                                self.ui.pressure_pv_c_5.value()]
        self.chart_pressure_c.append_data(group_c_values, group_c_press_values)

    def update_clock(self):
        self.ui.date_displ.setDateTime(QDateTime.currentDateTime())

    def disable_oil_group(self, channel, status):
        """
        Enable or disable UI controls during system operation.
        \nBật hoặc tắt các điều khiển UI trong quá trình hoạt động của máy.
        """
        if channel == "A":
            for i in range(5, 7):
                self.pressure_a_sv_obj[i].setEnabled(status)
            return

        elif channel == "B":
            for i in range(5, 7):
                self.pressure_b_sv_obj[i].setEnabled(status)
            return

        elif channel == "C":
            for i in range(5, 7):
                self.pressure_c_sv_obj[i].setEnabled(status)
            return

    def disable_pressure_group(self, channel, status):
        """
        Enable or disable UI controls during system operation.
        \nBật hoặc tắt các điều khiển UI trong quá trình hoạt động của máy.
        """
        if channel == "A":
            for i in range(1, 5):
                self.pressure_a_sv_obj[i].setEnabled(status)
            return

        elif channel == "B":
            for i in range(1, 5):
                self.pressure_b_sv_obj[i].setEnabled(status)
            return

        elif channel == "C":
            for i in range(1, 5):
                self.pressure_c_sv_obj[i].setEnabled(status)
            return

    def disable_heat_group(self, channel, status):
        """
        Enable or disable UI controls during system operation.
        \nBật hoặc tắt các điều khiển UI trong quá trình hoạt động của máy.
        """
        if channel == "A":  
            self.pressure_a_sv_obj[0].setEnabled(status)
            for i in [1]:
                self.temp_sv_obj[i].setEnabled(status)
                self.temp_h_alm_obj[i].setEnabled(status)
                self.temp_l_alm_obj[i].setEnabled(status)
            for i in range(1,4):
                self.temp_offset_obj[i].setEnabled(status)
            return

        elif channel == "B":
            self.pressure_b_sv_obj[0].setEnabled(status)
            for i in [2]:
                self.temp_sv_obj[i].setEnabled(status)
                self.temp_h_alm_obj[i].setEnabled(status)
                self.temp_l_alm_obj[i].setEnabled(status)
            for i in range(4,7):
                self.temp_offset_obj[i].setEnabled(status)
            return

        elif channel == "C":
            self.pressure_c_sv_obj[0].setEnabled(status)
            for i in [3]:
                self.temp_sv_obj[i].setEnabled(status)
                self.temp_h_alm_obj[i].setEnabled(status)
                self.temp_l_alm_obj[i].setEnabled(status)
            for i in range(7,10):
                self.temp_offset_obj[i].setEnabled(status)
            return

        elif channel == "T0":
            for i in [0]:
                self.temp_sv_obj[i].setEnabled(status)
                self.temp_h_alm_obj[i].setEnabled(status)
                self.temp_l_alm_obj[i].setEnabled(status)
                self.temp_offset_obj[i].setEnabled(status)
            return

    def add_row_to_list_history(self, product_name, group_name,
                                pressure_value, temp_value,
                                front_temp, mid_temp, end_temp):
        """Chỉ gom data vào pending"""
        try:
            if not hasattr(self, 'conn') or self.conn is None:
                return

            unit_suffix = "°F" if self._current_unit == 1 else "°C"
            def fmt(v):
                return f"{self.for_display_temp(v):.1f}{unit_suffix}"

            self._pending_rows.append([
                "", product_name, group_name,
                f"{pressure_value:.2f} bar",
                fmt(temp_value), fmt(front_temp), 
                fmt(mid_temp), fmt(end_temp),
                datetime.now().strftime("%H:%M - %d/%m/%Y")
            ])
        except Exception as e:
            self.logger.error(f"[add_row_to_list_history] Error: {e}")

    def _flush_history(self):
        """Mỗi 1s — batch insert DB + append vào table"""
        if not self._pending_rows or not self.conn:
            return

        rows = self._pending_rows.copy()
        self._pending_rows.clear()

        # DB insert trên thread riêng
        def _db_batch(rows=rows):
            try:
                self.conn.executemany('''
                    INSERT INTO history ("No.", "Name.", "Group.", "Pressure.", "T-Oven.",
                                    "Front.", "Middle.", "End.", "Date.")
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', rows)
                self.conn.commit()
            except Exception as e:
                self.logger.error(f"[_flush_history][DB] {e}")
        threading.Thread(target=_db_batch, daemon=True).start()

        # Append vào cache
        base_idx = len(self._all_rows_cache)
        for i, row_data in enumerate(rows):
            row_data[0] = str(base_idx + i + 1)
            self._all_rows_cache.append(row_data)

        # Append vào table (rows mới nhất luôn hiển thị ở bottom)
        table = self.ui.list_history
        scrollbar = table.verticalScrollBar()
        was_at_bottom = scrollbar.value() >= scrollbar.maximum() - 5

        table.setUpdatesEnabled(False)
        for row_data in rows:
            row_pos = table.rowCount()
            table.insertRow(row_pos)
            for col, value in enumerate(row_data):
                item = QTableWidgetItem(str(value))
                if col == 0:
                    item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row_pos, col, item)

            if table.rowCount() > self._table_display:
                table.removeRow(0)

        table.setUpdatesEnabled(True)

        if was_at_bottom:
            table.scrollToBottom()

    def auto_export_all_tables_to_excel(self):
        """
        Export both the list_history and list_err tables to two dif Excel files
        \nExport cả bảng list_history và list_err ra 2 file Excel riêng biệt.
        """
        current_date = datetime.now().strftime("%d_%m_%Y")

        file_export = self.ui.list_history
        default_filename_done = f"History {current_date}.xlsx"
        list_history_folder = Path(self.stk_mch_folder) / "DB_address"
        filename_done = list_history_folder / default_filename_done
        if filename_done:
            self._export_table_to_excel(file_export, str(filename_done))
            self.logger.info("Autosave Done")

    def save_after_quit_export_all_tables_to_excel(self):
        """
        Export both the list_history and list_err tables to two dif Excel files
        \nExport cả bảng list_history và list_err ra 2 file Excel riêng biệt.
        """
        current_date = datetime.now().strftime("%d_%m_%Y")

        file_export = self.ui.list_history
        default_filename_done = f"History {current_date}.xlsx"
        list_history_folder = Path(self.stk_mch_folder) / "Excel"
        filename_done = list_history_folder / default_filename_done
        if filename_done:
            self._export_table_to_excel(file_export, str(filename_done))
            self.logger.info("Save Done")

    def export_all_tables_to_excel_btn(self):
        """
        Export both the list_history and list_err tables to two dif Excel files
        \nExport cả bảng list_history và list_err ra 2 file Excel riêng biệt.
        """
        current_date = datetime.now().strftime("%d_%m_%Y")

        file_export = self.ui.list_history_2 if self.ui.stackedWidget_4.currentIndex() == 1 else self.ui.list_history
        default_filename_done = f"History {current_date}.xlsx"
        list_history_folder = Path(self.stk_mch_folder) / "Excel"
        filename_done = list_history_folder / default_filename_done
        if filename_done:
            self.export_table_to_excel(file_export, str(filename_done))

    def _export_table_to_excel(self, table_widget, filename):
        # Lấy headers từ QTableWidget
        headers = []
        for col in range(table_widget.columnCount()):
            header = table_widget.horizontalHeaderItem(col)
            headers.append(header.text() if header else f"Column {col + 1}")

        # Lấy data từ QTableWidget
        rows = []
        for row in range(table_widget.rowCount()):
            row_data = []
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                row_data.append(item.text() if item else "")
            rows.append(row_data)

        # Tạo DataFrame và xuất Excel
        df = pd.DataFrame(rows, columns=headers)
        
        try:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Sheet1")
                
                # Auto-fit column width
                worksheet = writer.sheets["Sheet1"]
                for col_idx, col_name in enumerate(df.columns, start=1):
                    max_len = max(
                        len(str(col_name)),
                        df.iloc[:, col_idx - 1].astype(str).map(len).max() if len(df) > 0 else 0
                    )
                    worksheet.column_dimensions[
                        worksheet.cell(1, col_idx).column_letter
                    ].width = max_len + 4
                header_font = Font(sz=12, bold=True, color="1E40AF")
                border_bottom = Border(bottom=Side(style="medium", color="3B82F6"))
                for col_idx in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.border = border_bottom
        except Exception as e:
            self.logger.error(f"[Main]-[export_table_to_excel]: Export file Error: {e}")
            # ltmessage.error(self, "Export Error", str(e))

    def export_table_to_excel(self, table: QTableWidget, default_filename: str = "export"):
        """Xuất QTableWidget ra file Excel"""
        
        # Chọn nơi lưu file
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            default_filename,
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:  # User cancel
            return
        
        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"

        # Lấy headers từ QTableWidget
        headers = []
        for col in range(table.columnCount()):
            header = table.horizontalHeaderItem(col)
            headers.append(header.text() if header else f"Column {col + 1}")
        
        # Lấy data từ QTableWidget
        rows = []
        for row in range(table.rowCount()):
            row_data = []
            for col in range(table.columnCount()):
                item = table.item(row, col)
                row_data.append(item.text() if item else "")
            rows.append(row_data)

        # Tạo DataFrame và xuất Excel
        df = pd.DataFrame(rows, columns=headers)
        
        try:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Sheet1")
                
                # Auto-fit column width
                worksheet = writer.sheets["Sheet1"]
                for col_idx, col_name in enumerate(df.columns, start=1):
                    max_len = max(
                        len(str(col_name)),
                        df.iloc[:, col_idx - 1].astype(str).map(len).max() if len(df) > 0 else 0
                    )
                    worksheet.column_dimensions[
                        worksheet.cell(1, col_idx).column_letter
                    ].width = max_len + 4

                header_font = Font(bold=True, color="1E40AF")
                border_bottom = Border(bottom=Side(style="medium", color="3B82F6"))
                for col_idx in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.border = border_bottom
            ltmessage.information(self, "Success", f"Exported successfully:\n{file_path}")

        except Exception as e:
            ltmessage.error(self, "Error", f"Export failed:\n{e}")
            
    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, '_maximized_chart_idx') and self._maximized_chart_idx == -1:
            QTimer.singleShot(50, self._save_grid_rects)
            
    def closeEvent(self, event):
        reply = ltmessage.question(
            self, "Exit Confirmation", "Are you sure you want to exit?"
        )

        if reply == ltmessage.Yes:
            self.logger.info("Application is closing...")
            try:
                self._close_event_cleanup()
            except Exception as e:
                ltmessage.error(self, "Error", f"Failed to cleanup: {e}")
                event.ignore()
            event.accept()
        else:
            event.ignore()
        
    def _close_event_cleanup(self):
        self._cleanup()

    def _cleanup(self):
        for timer in getattr(self, 'all_timer', []):
            if timer.isActive():
                timer.stop()
            timer.deleteLater()
        self.all_timer.clear()
        self.hide()

        if self.plc_read_worker:
            self.plc_read_worker.stop()
        if self.plc_writer_worker:
            self.plc_writer_worker.stop()

        if self.plc_read_thread:
            self.plc_read_thread.wait()
        if self.plc_writer_thread:
            self.plc_writer_thread.wait()
        
        self.stop_simulate_threads() if SIMULATE else None

    def stop_simulate_threads(self):
        try:
            simulate_thread = self.thread_dict.get("data_simulator")

            if simulate_thread is not None:

                # Request stop flag nếu có
                if hasattr(simulate_thread, "stop"):
                    simulate_thread.stop()

                # Stop event loop của thread
                simulate_thread.quit()

                # Chờ thread tắt
                if not simulate_thread.wait(3000):
                    print("Force terminate DataSimulator thread")
                    simulate_thread.terminate()
                    simulate_thread.wait()

                print("DataSimulator stopped")

                # Xóa khỏi dict
                del self.thread_dict["data_simulator"]

        except Exception as e:
            print(f"Stop thread error: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    window_id = 'StrikeMachine_Instance'
    shared_mem_id = 'StrikeMachine_SharedMem'
    
    semaphore = QSystemSemaphore(window_id, 1)
    semaphore.acquire()
    
    if sys.platform != 'win32':
        nix_fix_shared_mem = QSharedMemory(shared_mem_id)
        if nix_fix_shared_mem.attach():
            nix_fix_shared_mem.detach()
    
    shared_memory = QSharedMemory(shared_mem_id)
    
    if shared_memory.attach():
        is_running = True
    else:
        shared_memory.create(1)
        is_running = False
    
    semaphore.release()
    
    if is_running:
        sys.exit(1)

    window = StrikeMachine()

    def center_window(win):
        screen = QApplication.primaryScreen().availableGeometry()
        x = (screen.width() - win.width()) // 2
        y = (screen.height() - win.height()) // 2
        win.move(x, y)

    center_window(window)
    window.show()

    def cleanup():
        try:
            # window.auto_export_all_tables_to_excel()
            window.save_after_quit_export_all_tables_to_excel()
            window._close_event_cleanup()
        except:
            pass

    atexit.register(cleanup)
    
    sys.exit(app.exec())